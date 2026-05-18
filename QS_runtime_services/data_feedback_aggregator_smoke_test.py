"""
data_feedback_aggregator.py 的端到端冒烟测试。

本脚本会在专用测试目录里写入 user_data、usage、feedback 的 JSON 小文件，
然后调用汇总逻辑生成并续写 summary Excel。默认不会访问真实 Public 或共享目录。

运行方式：
    python -m QS_runtime_services.data_feedback_aggregator_smoke_test
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
import json
from pathlib import Path
import shutil
import sys
from typing import Any

from openpyxl import load_workbook

try:
    from .data_feedback_aggregator import _aggregate_inbox
except ImportError:
    ROOT_DIR = Path(__file__).resolve().parent.parent
    if str(ROOT_DIR) not in sys.path:
        sys.path.insert(0, str(ROOT_DIR))
    from QS_runtime_services.data_feedback_aggregator import _aggregate_inbox


DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "_aggregator_test_output"

USER_DATA_HEADERS = [
    "填写时间", "主机名", "用户姓名", "部门", "LM大组", "PL小组", "应用项目名",
    "App版本",
]
USAGE_HEADERS = ["主机名", "启动时间", "关闭时间"]
FEEDBACK_HEADERS = [
    "提交时间", "用户姓名", "主机名", "部门", "LM大组", "PL小组", "应用项目名",
    "使用强度", "单点效率提高程度", "单点效率补充", "整体效率提高程度",
    "整体效率补充", "需求重要度", "需求紧急度", "需求维度", "需求描述",
    "附件", "App版本",
]


@dataclass(frozen=True)
class TestPaths:
    root: Path
    data_feedback: Path
    inbox: Path
    user_data_summary: Path
    usage_summary: Path
    feedback_summary: Path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="构造 JSON 小文件并验证 data_feedback_aggregator 汇总结果。"
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="测试输出目录；默认 QS_runtime_services/_aggregator_test_output。",
    )
    parser.add_argument(
        "--no-clean",
        action="store_true",
        help="不清理旧测试输出，适合手动观察续写行为。",
    )
    args = parser.parse_args(argv)

    paths = _make_paths(Path(args.output_dir))
    if not args.no_clean:
        _clean_output_dir(paths.root)
    paths.inbox.mkdir(parents=True, exist_ok=True)

    print(f"测试目录: {paths.root}")
    print("写入第一批 JSON：首次生成、重复 user_data 去重、缺失 user_data 主机。")
    _write_batch_one(paths.inbox)
    stats1 = _aggregate(paths)
    _assert_stats("第一批", stats1, processed=5, deleted=5, failed=0)
    _assert_after_batch_one(paths)

    print("写入第二批 JSON：续写、用户资料更新、补齐缺失资料、保留异常 JSON。")
    _write_batch_two(paths.inbox)
    stats2 = _aggregate(paths)
    _assert_stats("第二批", stats2, processed=6, deleted=6, failed=2)
    _assert_after_batch_two(paths)

    remaining_json = sorted(path.name for path in paths.inbox.glob("*.json"))
    _assert_equal(
        remaining_json,
        ["bad_malformed_json.json", "bad_unknown_sheet.json"],
        "异常 JSON 应被保留，正常 JSON 应被删除",
    )

    print("测试通过。生成文件：")
    print(f"- {paths.user_data_summary}")
    print(f"- {paths.usage_summary}")
    print(f"- {paths.feedback_summary}")
    print(f"异常样本保留在: {paths.inbox}")
    return 0


def _make_paths(root: Path) -> TestPaths:
    data_feedback = root / "data_feedback"
    return TestPaths(
        root=root,
        data_feedback=data_feedback,
        inbox=data_feedback / "inbox",
        user_data_summary=data_feedback / "Quick_Sparam_user_data_summary.xlsx",
        usage_summary=data_feedback / "Quick_Sparam_usage_summary.xlsx",
        feedback_summary=data_feedback / "Quick_Sparam_feedback_summary.xlsx",
    )


def _clean_output_dir(root: Path) -> None:
    resolved_root = root.resolve()
    repo_root = Path(__file__).resolve().parent.parent.resolve()
    if root.exists():
        if not _is_relative_to(resolved_root, repo_root):
            raise RuntimeError(f"拒绝清理仓库外目录: {resolved_root}")
        if "aggregator_test_output" not in root.name:
            raise RuntimeError(
                "拒绝清理非专用测试目录；请使用 --no-clean，或把目录名设置为 "
                "aggregator_test_output。"
            )
        shutil.rmtree(root)


def _write_batch_one(inbox: Path) -> None:
    user_alpha = [
        "2026-05-12T08:50:00+08:00",
        "HOST-ALPHA",
        "张三",
        "部门A",
        "大组A",
        "小组A1",
        "项目甲",
        "2026.03",
    ]
    _write_payload(
        inbox,
        "user_data_alpha_1.json",
        "用户信息",
        USER_DATA_HEADERS,
        user_alpha,
        "user_data_alpha_v1",
    )
    _write_payload(
        inbox,
        "user_data_alpha_1_duplicate.json",
        "用户信息",
        USER_DATA_HEADERS,
        user_alpha,
        "user_data_alpha_v1",
    )
    _write_payload(
        inbox,
        "usage_alpha_1.json",
        "使用记录",
        USAGE_HEADERS,
        ["HOST-ALPHA", "2026-05-12T09:00:00+08:00", "2026-05-12T09:30:00+08:00"],
        "usage_alpha_1",
    )
    _write_payload(
        inbox,
        "usage_beta_missing_profile_1.json",
        "使用记录",
        USAGE_HEADERS,
        ["HOST-BETA", "2026-05-12T11:00:00+08:00", "2026-05-12T11:10:00+08:00"],
        "usage_beta_1",
    )
    _write_payload(
        inbox,
        "feedback_alpha_1.json",
        "评价反馈",
        FEEDBACK_HEADERS,
        [
            "2026-05-12T09:40:00+08:00",
            "张三",
            "HOST-ALPHA",
            "部门A",
            "大组A",
            "小组A1",
            "项目甲",
            "2-3h/周",
            "50%-75%",
            "减少手动整理端口时间",
            "25%-50%",
            "",
            "高",
            "较急",
            "功能改进",
            "希望增加批量模板。",
            "",
            "2026.03",
        ],
        "feedback_alpha_1",
    )


def _write_batch_two(inbox: Path) -> None:
    _write_payload(
        inbox,
        "user_data_alpha_2_update.json",
        "用户信息",
        USER_DATA_HEADERS,
        [
            "2026-05-12T12:00:00+08:00",
            "HOST-ALPHA",
            "张三更新",
            "部门A",
            "大组A",
            "小组A2",
            "项目甲,项目乙",
            "2026.03",
        ],
        "user_data_alpha_v2",
    )
    _write_payload(
        inbox,
        "user_data_beta_1.json",
        "用户信息",
        USER_DATA_HEADERS,
        [
            "2026-05-12T12:30:00+08:00",
            "HOST-BETA",
            "李四",
            "部门B",
            "大组B",
            "小组B1",
            "项目乙",
            "2026.03",
        ],
        "user_data_beta_v1",
    )
    _write_payload(
        inbox,
        "usage_alpha_2.json",
        "使用记录",
        USAGE_HEADERS,
        ["HOST-ALPHA", "2026-05-12T10:00:00+08:00", "2026-05-12T10:45:00+08:00"],
        "usage_alpha_2",
    )
    _write_payload(
        inbox,
        "usage_beta_2.json",
        "使用记录",
        USAGE_HEADERS,
        ["HOST-BETA", "2026-05-12T12:00:00+08:00", "2026-05-12T12:20:00+08:00"],
        "usage_beta_2",
    )
    _write_payload(
        inbox,
        "usage_gamma_missing_profile_1.json",
        "使用记录",
        USAGE_HEADERS,
        ["HOST-GAMMA", "2026-05-12T13:00:00+08:00", "2026-05-12T13:05:00+08:00"],
        "usage_gamma_1",
    )
    _write_payload(
        inbox,
        "feedback_beta_1.json",
        "评价反馈",
        FEEDBACK_HEADERS,
        [
            "2026-05-12T12:35:00+08:00",
            "李四",
            "HOST-BETA",
            "部门B",
            "大组B",
            "小组B1",
            "项目乙",
            "",
            "",
            "",
            "",
            "",
            "中",
            "一般",
            "Bug反馈",
            "只填写需求文本也应允许提交。",
            "",
            "2026.03",
        ],
        "feedback_beta_1",
    )
    _write_payload(
        inbox,
        "bad_unknown_sheet.json",
        "未知表",
        ["字段"],
        ["无法识别的 sheet_name 应保留"],
        "bad_unknown_sheet",
    )
    (inbox / "bad_malformed_json.json").write_text(
        "{ this is not valid json", encoding="utf-8"
    )


def _write_payload(
    inbox: Path,
    filename: str,
    sheet_name: str,
    headers: list[str],
    row: list[Any],
    record_id: str,
) -> None:
    record = {
        str(header): row[index] if index < len(row) else ""
        for index, header in enumerate(headers)
    }
    payload = {
        "schema_version": 1,
        "record_id": record_id,
        "record_type": _record_type(sheet_name),
        "record_name": sheet_name,
        "sheet_name": sheet_name,
        "headers": headers,
        "row": row,
        "record": record,
        "created_at": "2026-05-12T14:00:00+08:00",
    }
    (inbox / filename).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _record_type(sheet_name: str) -> str:
    if sheet_name == "用户信息":
        return "user_data"
    if sheet_name == "评价反馈":
        return "feedback"
    if sheet_name == "使用记录":
        return "usage"
    return "unknown"


def _aggregate(paths: TestPaths) -> dict[str, int]:
    messages: list[str] = []
    stats = _aggregate_inbox(
        target_name="测试",
        inbox_dir=paths.inbox,
        user_data_summary_paths=[paths.user_data_summary],
        usage_summary_paths=[paths.usage_summary],
        feedback_summary_paths=[paths.feedback_summary],
        printer=messages.append,
    )
    for message in messages:
        print(f"  {message}")
    return stats


def _assert_after_batch_one(paths: TestPaths) -> None:
    user_rows = _sheet_records(paths.user_data_summary, "用户信息")
    usage_rows = _sheet_records(paths.usage_summary, "使用记录")
    feedback_rows = _sheet_records(paths.feedback_summary, "评价反馈")
    host_rows = _sheet_records(paths.usage_summary, "主机累计统计")

    _assert_equal(len(user_rows), 1, "第一批重复 user_data 应只写入一条")
    _assert_equal(len(usage_rows), 2, "第一批 usage 应写入两条")
    _assert_equal(len(feedback_rows), 1, "第一批 feedback 应写入一条")

    alpha = _row_by_value(host_rows, "主机名", "HOST-ALPHA")
    beta = _row_by_value(host_rows, "主机名", "HOST-BETA")
    _assert_equal(alpha["用户姓名"], "张三", "HOST-ALPHA 应填充用户姓名")
    _assert_equal(alpha["使用次数"], 1, "HOST-ALPHA 第一批使用次数")
    _assert_equal(alpha["总使用秒数"], 1800, "HOST-ALPHA 第一批累计秒数")
    _assert_equal(beta["用户姓名"], None, "缺失 user_data 的 HOST-BETA 用户姓名应为空")
    _assert_equal(beta["部门"], None, "缺失 user_data 的 HOST-BETA 部门应为空")
    _assert_equal(beta["总使用秒数"], 600, "HOST-BETA 第一批累计秒数")

    _assert_workbook_format(paths.user_data_summary)
    _assert_workbook_format(paths.usage_summary)
    _assert_workbook_format(paths.feedback_summary)


def _assert_after_batch_two(paths: TestPaths) -> None:
    user_rows = _sheet_records(paths.user_data_summary, "用户信息")
    usage_rows = _sheet_records(paths.usage_summary, "使用记录")
    feedback_rows = _sheet_records(paths.feedback_summary, "评价反馈")
    host_rows = _sheet_records(paths.usage_summary, "主机累计统计")

    _assert_equal(len(user_rows), 3, "续写后 user_data 历史记录数量")
    _assert_equal(len(usage_rows), 5, "续写后 usage 历史记录数量")
    _assert_equal(len(feedback_rows), 2, "续写后 feedback 历史记录数量")

    alpha = _row_by_value(host_rows, "主机名", "HOST-ALPHA")
    beta = _row_by_value(host_rows, "主机名", "HOST-BETA")
    gamma = _row_by_value(host_rows, "主机名", "HOST-GAMMA")

    _assert_equal(alpha["用户姓名"], "张三更新", "HOST-ALPHA 应使用最新用户资料")
    _assert_equal(alpha["PL小组"], "小组A2", "HOST-ALPHA 应使用最新 PL 小组")
    _assert_equal(alpha["使用次数"], 2, "HOST-ALPHA 续写后使用次数")
    _assert_equal(alpha["总使用秒数"], 4500, "HOST-ALPHA 续写后累计秒数")

    _assert_equal(beta["用户姓名"], "李四", "HOST-BETA 第二批补齐用户姓名")
    _assert_equal(beta["部门"], "部门B", "HOST-BETA 第二批补齐部门")
    _assert_equal(beta["使用次数"], 2, "HOST-BETA 续写后使用次数")
    _assert_equal(beta["总使用秒数"], 1800, "HOST-BETA 续写后累计秒数")

    _assert_equal(gamma["用户姓名"], None, "HOST-GAMMA 无 user_data 时用户姓名应为空")
    _assert_equal(gamma["使用次数"], 1, "HOST-GAMMA 使用次数")
    _assert_equal(gamma["总使用秒数"], 300, "HOST-GAMMA 累计秒数")

    _assert_no_repeated_header_rows(paths.user_data_summary, "用户信息")
    _assert_no_repeated_header_rows(paths.usage_summary, "使用记录")
    _assert_no_repeated_header_rows(paths.usage_summary, "主机累计统计")
    _assert_no_repeated_header_rows(paths.feedback_summary, "评价反馈")
    _assert_workbook_format(paths.user_data_summary)
    _assert_workbook_format(paths.usage_summary)
    _assert_workbook_format(paths.feedback_summary)


def _assert_stats(
    label: str, stats: dict[str, int], processed: int, deleted: int, failed: int
) -> None:
    _assert_equal(stats.get("processed"), processed, f"{label} processed 数量")
    _assert_equal(stats.get("deleted"), deleted, f"{label} deleted 数量")
    _assert_equal(stats.get("failed"), failed, f"{label} failed 数量")


def _sheet_records(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise AssertionError(f"{workbook_path} 缺少 sheet: {sheet_name}")
    worksheet = workbook[sheet_name]
    headers = [
        worksheet.cell(row=1, column=column_index).value
        for column_index in range(1, worksheet.max_column + 1)
    ]
    records: list[dict[str, Any]] = []
    for row_index in range(2, worksheet.max_row + 1):
        record = {
            header: worksheet.cell(row=row_index, column=column_index).value
            for column_index, header in enumerate(headers, start=1)
            if header
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def _row_by_value(rows: list[dict[str, Any]], key: str, value: Any) -> dict[str, Any]:
    for row in rows:
        if row.get(key) == value:
            return row
    raise AssertionError(f"没有找到 {key}={value} 的记录")


def _assert_no_repeated_header_rows(workbook_path: Path, sheet_name: str) -> None:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[sheet_name]
    headers = [
        worksheet.cell(row=1, column=column_index).value
        for column_index in range(1, worksheet.max_column + 1)
    ]
    for row_index in range(2, worksheet.max_row + 1):
        row_values = [
            worksheet.cell(row=row_index, column=column_index).value
            for column_index in range(1, len(headers) + 1)
        ]
        if row_values == headers:
            raise AssertionError(f"{sheet_name} 第 {row_index} 行重复出现表头")


def _assert_workbook_format(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    for worksheet in workbook.worksheets:
        _assert_equal(worksheet.freeze_panes, "A2", f"{worksheet.title} 冻结首行")
        _assert(
            bool(worksheet.auto_filter.ref),
            f"{worksheet.title} 应设置自动筛选范围",
        )
        first_width = worksheet.column_dimensions["A"].width or 0
        _assert(
            first_width >= 10,
            f"{worksheet.title} A 列列宽应不小于 10，当前 {first_width}",
        )


def _assert_equal(actual: Any, expected: Any, message: str) -> None:
    if actual != expected:
        raise AssertionError(f"{message}: 期望 {expected!r}，实际 {actual!r}")


def _assert(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
        return True
    except ValueError:
        return False


if __name__ == "__main__":
    raise SystemExit(main())
