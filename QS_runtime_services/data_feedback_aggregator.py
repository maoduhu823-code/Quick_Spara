"""
data_feedback/inbox 记录汇总程序。

用法：
    python -m QS_runtime_services.data_feedback_aggregator --once
    python -m QS_runtime_services.data_feedback_aggregator --target developer --once
    python -m QS_runtime_services.data_feedback_aggregator --target developer --interval-seconds 600

程序会读取 inbox 下的独立 JSON 记录，写入 summary Excel：
- Quick_Sparam_user_data_summary.xlsx
- Quick_Sparam_usage_summary.xlsx
- Quick_Sparam_feedback_summary.xlsx

只有 summary 保存成功后，才会删除对应 JSON 小文件。
"""

from __future__ import annotations

import argparse
from contextlib import contextmanager
from datetime import datetime
import json
import os
from pathlib import Path
import sys
import time
from typing import Any, Iterator

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook

try:
    from .path_config import (
        DEVELOPER_RECORD_INBOX_DIRS_BY_PLATFORM,
        FEEDBACK_DEVELOPER_SUMMARY_PATHS_BY_PLATFORM,
        FEEDBACK_LOCAL_SUMMARY_PATHS_BY_PLATFORM,
        LOCAL_RECORD_INBOX_DIRS_BY_PLATFORM,
        USER_DATA_DEVELOPER_SUMMARY_PATHS_BY_PLATFORM,
        USER_DATA_LOCAL_SUMMARY_PATHS_BY_PLATFORM,
        USAGE_DEVELOPER_SUMMARY_PATHS_BY_PLATFORM,
        USAGE_LOCAL_SUMMARY_PATHS_BY_PLATFORM,
    )
    from .record_writer import configured_paths
except ImportError:
    ROOT_DIR = Path(__file__).resolve().parent.parent
    if str(ROOT_DIR) not in sys.path:
        sys.path.insert(0, str(ROOT_DIR))
    from QS_runtime_services.path_config import (
        DEVELOPER_RECORD_INBOX_DIRS_BY_PLATFORM,
        FEEDBACK_DEVELOPER_SUMMARY_PATHS_BY_PLATFORM,
        FEEDBACK_LOCAL_SUMMARY_PATHS_BY_PLATFORM,
        LOCAL_RECORD_INBOX_DIRS_BY_PLATFORM,
        USER_DATA_DEVELOPER_SUMMARY_PATHS_BY_PLATFORM,
        USER_DATA_LOCAL_SUMMARY_PATHS_BY_PLATFORM,
        USAGE_DEVELOPER_SUMMARY_PATHS_BY_PLATFORM,
        USAGE_LOCAL_SUMMARY_PATHS_BY_PLATFORM,
    )
    from QS_runtime_services.record_writer import configured_paths


META_HEADERS = ["记录ID", "源文件名", "汇总时间"]
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 42


def aggregate_records(target: str = "all", printer: Any = print) -> dict[str, int]:
    """汇总指定目标的 inbox。target: all/local/developer。"""
    stats = {"processed": 0, "deleted": 0, "failed": 0}
    for target_config in _target_configs(target):
        for inbox_dir in target_config["inbox_dirs"]:
            result = _aggregate_inbox(
                target_name=target_config["name"],
                inbox_dir=inbox_dir,
                user_data_summary_paths=target_config["user_data_summary_paths"],
                usage_summary_paths=target_config["usage_summary_paths"],
                feedback_summary_paths=target_config["feedback_summary_paths"],
                printer=printer,
            )
            for key, value in result.items():
                stats[key] = stats.get(key, 0) + value
    return stats


def _target_configs(target: str) -> list[dict[str, Any]]:
    app_base_dir = _app_base_dir()
    configs = {
        "local": {
            "name": "本地",
            "inbox_dirs": configured_paths(
                LOCAL_RECORD_INBOX_DIRS_BY_PLATFORM, app_base_dir
            ),
            "usage_summary_paths": configured_paths(
                USAGE_LOCAL_SUMMARY_PATHS_BY_PLATFORM, app_base_dir
            ),
            "user_data_summary_paths": configured_paths(
                USER_DATA_LOCAL_SUMMARY_PATHS_BY_PLATFORM, app_base_dir
            ),
            "feedback_summary_paths": configured_paths(
                FEEDBACK_LOCAL_SUMMARY_PATHS_BY_PLATFORM, app_base_dir
            ),
        },
        "developer": {
            "name": "开发者",
            "inbox_dirs": configured_paths(
                DEVELOPER_RECORD_INBOX_DIRS_BY_PLATFORM, app_base_dir
            ),
            "usage_summary_paths": configured_paths(
                USAGE_DEVELOPER_SUMMARY_PATHS_BY_PLATFORM, app_base_dir
            ),
            "user_data_summary_paths": configured_paths(
                USER_DATA_DEVELOPER_SUMMARY_PATHS_BY_PLATFORM, app_base_dir
            ),
            "feedback_summary_paths": configured_paths(
                FEEDBACK_DEVELOPER_SUMMARY_PATHS_BY_PLATFORM, app_base_dir
            ),
        },
    }
    if target == "all":
        return [configs["local"], configs["developer"]]
    return [configs[target]]


def _aggregate_inbox(
    target_name: str,
    inbox_dir: Path,
    user_data_summary_paths: list[Path],
    usage_summary_paths: list[Path],
    feedback_summary_paths: list[Path],
    printer: Any,
) -> dict[str, int]:
    stats = {"processed": 0, "deleted": 0, "failed": 0}
    if not inbox_dir.exists():
        printer(f"{target_name} inbox 不存在，跳过: {inbox_dir}")
        return stats

    with _inbox_lock(inbox_dir) as locked:
        if not locked:
            printer(f"{target_name} inbox 正在被其他汇总任务处理，跳过: {inbox_dir}")
            return stats

        payload_groups: dict[str, list[tuple[Path, dict[str, Any]]]] = {
            "user_data": [],
            "usage": [],
            "feedback": [],
        }
        for json_path in sorted(inbox_dir.glob("*.json")):
            try:
                payload = _load_payload(json_path)
                kind = _payload_kind(payload)
            except Exception as exc:
                stats["failed"] += 1
                printer(f"记录文件读取失败，保留待排查: {json_path}，错误: {exc}")
                continue
            if kind not in payload_groups:
                stats["failed"] += 1
                printer(f"无法识别记录类型，保留待排查: {json_path}")
                continue
            payload_groups[kind].append((json_path, payload))

        user_data_count, user_data_saved_path = _write_group_and_delete(
            "用户信息",
            payload_groups["user_data"],
            user_data_summary_paths,
            printer,
            stats,
        )
        usage_count, usage_saved_path = _write_group_and_delete(
            "使用记录",
            payload_groups["usage"],
            usage_summary_paths,
            printer,
            stats,
        )
        feedback_count, _feedback_saved_path = _write_group_and_delete(
            "评价反馈",
            payload_groups["feedback"],
            feedback_summary_paths,
            printer,
            stats,
        )
        stats["processed"] += user_data_count + usage_count + feedback_count
        if user_data_count or usage_count:
            _rebuild_host_usage_summary(
                _first_existing_or_first(usage_summary_paths, usage_saved_path),
                _first_existing_or_first(user_data_summary_paths, user_data_saved_path),
                printer,
            )
    return stats


def _write_group_and_delete(
    group_name: str,
    items: list[tuple[Path, dict[str, Any]]],
    summary_paths: list[Path],
    printer: Any,
    stats: dict[str, int],
) -> tuple[int, Path | None]:
    if not items:
        return 0, None
    if not summary_paths:
        stats["failed"] += len(items)
        printer(f"{group_name}没有配置 summary 输出路径，保留 {len(items)} 个 JSON。")
        return 0, None

    try:
        workbook = _load_or_create_workbook(summary_paths)
        for json_path, payload in items:
            _append_payload(workbook, payload, json_path)
        saved_path = _save_workbook(workbook, summary_paths)
    except Exception as exc:
        stats["failed"] += len(items)
        printer(f"{group_name}汇总失败，保留 JSON 待下次重试。错误: {exc}")
        return 0, None

    deleted = 0
    for json_path, _payload in items:
        try:
            json_path.unlink()
            deleted += 1
        except OSError as exc:
            stats["failed"] += 1
            printer(f"汇总已写入但删除记录文件失败: {json_path}，错误: {exc}")
    stats["deleted"] += deleted
    printer(f"{group_name}已汇总 {len(items)} 条到 {saved_path}，删除 {deleted} 个 JSON。")
    return len(items), saved_path


def _rebuild_host_usage_summary(
    usage_summary_path: Path | None,
    user_data_summary_path: Path | None,
    printer: Any,
) -> None:
    if usage_summary_path is None or not usage_summary_path.exists():
        return

    try:
        workbook = load_workbook(usage_summary_path)
        if "使用记录" not in workbook.sheetnames:
            return
        records = _read_sheet_records(workbook["使用记录"])
        profiles = _load_latest_user_profiles(user_data_summary_path)
        aggregates = _aggregate_usage_by_host(records)

        if "主机累计统计" in workbook.sheetnames:
            del workbook["主机累计统计"]
        worksheet = workbook.create_sheet("主机累计统计")
        headers = [
            "主机名", "用户姓名", "部门", "LM大组", "PL小组", "应用项目名",
            "使用次数", "总使用秒数", "总使用分钟", "总使用时长",
            "首次启动时间", "最近关闭时间",
        ]
        worksheet.append(headers)

        for host_name in sorted(aggregates):
            item = aggregates[host_name]
            profile = profiles.get(host_name, {})
            worksheet.append([
                host_name,
                profile.get("用户姓名", ""),
                profile.get("部门", ""),
                profile.get("LM大组", ""),
                profile.get("PL小组", ""),
                profile.get("应用项目名", ""),
                item["count"],
                item["total_seconds"],
                round(item["total_seconds"] / 60, 2),
                _format_duration(item["total_seconds"]),
                _format_datetime(item["first_start"]),
                _format_datetime(item["last_end"]),
            ])

        _format_workbook(workbook)
        workbook.save(usage_summary_path)
    except Exception as exc:
        printer(f"主机累计统计刷新失败，原始记录已保留。错误: {exc}")


def _aggregate_usage_by_host(records: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    aggregates: dict[str, dict[str, Any]] = {}
    for record in records:
        host_name = str(record.get("主机名", "")).strip()
        if not host_name:
            continue
        start_dt = _parse_datetime(record.get("启动时间"))
        end_dt = _parse_datetime(record.get("关闭时间"))
        if start_dt is None or end_dt is None:
            continue
        elapsed_seconds = max(0, int(round((end_dt - start_dt).total_seconds())))

        item = aggregates.setdefault(
            host_name,
            {
                "count": 0,
                "total_seconds": 0,
                "first_start": start_dt,
                "last_end": end_dt,
            },
        )
        item["count"] += 1
        item["total_seconds"] += elapsed_seconds
        if start_dt < item["first_start"]:
            item["first_start"] = start_dt
        if end_dt > item["last_end"]:
            item["last_end"] = end_dt
    return aggregates


def _load_latest_user_profiles(user_data_summary_path: Path | None) -> dict[str, dict[str, Any]]:
    if user_data_summary_path is None or not user_data_summary_path.exists():
        return {}
    workbook = load_workbook(user_data_summary_path, read_only=True, data_only=True)
    if "用户信息" not in workbook.sheetnames:
        return {}

    profiles: dict[str, dict[str, Any]] = {}
    profile_times: dict[str, float] = {}
    for record in _read_sheet_records(workbook["用户信息"]):
        host_name = str(record.get("主机名", "")).strip()
        if not host_name:
            continue
        updated_at = _parse_datetime(record.get("填写时间"))
        updated_key = updated_at.timestamp() if updated_at is not None else 0.0
        if host_name not in profiles or updated_key >= profile_times[host_name]:
            profiles[host_name] = record
            profile_times[host_name] = updated_key
    return profiles


def _read_sheet_records(worksheet: Any) -> list[dict[str, Any]]:
    headers = [
        str(worksheet.cell(row=1, column=column_index).value or "")
        for column_index in range(1, worksheet.max_column + 1)
    ]
    records: list[dict[str, Any]] = []
    for row_index in range(2, worksheet.max_row + 1):
        record = {
            header: worksheet.cell(row=row_index, column=column_index).value
            for column_index, header in enumerate(headers, start=1)
            if header
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def _parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


def _format_datetime(value: datetime) -> str:
    return value.isoformat(timespec="seconds")


def _format_duration(seconds: int) -> str:
    hours, remainder = divmod(int(seconds), 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours}小时{minutes}分钟{secs}秒"
    if minutes:
        return f"{minutes}分钟{secs}秒"
    return f"{secs}秒"


def _first_existing_or_first(paths: list[Path], preferred: Path | None = None) -> Path | None:
    if preferred is not None:
        return preferred
    for path in paths:
        if path.exists():
            return path
    return paths[0] if paths else None


def _load_payload(json_path: Path) -> dict[str, Any]:
    with json_path.open("r", encoding="utf-8-sig") as f:
        payload = json.load(f)
    if not isinstance(payload, dict):
        raise ValueError("JSON 顶层不是对象")
    if not payload.get("sheet_name") or not payload.get("record"):
        raise ValueError("缺少 sheet_name 或 record")
    return payload


def _payload_kind(payload: dict[str, Any]) -> str:
    sheet_name = str(payload.get("sheet_name", ""))
    if sheet_name in {"用户信息", "用户问卷"}:
        return "user_data"
    if sheet_name == "使用记录":
        return "usage"
    if sheet_name == "评价反馈":
        return "feedback"
    return ""


def _load_or_create_workbook(summary_paths: list[Path]) -> Workbook:
    for path in summary_paths:
        if path.exists():
            return load_workbook(path)
    return Workbook()


def _append_payload(workbook: Workbook, payload: dict[str, Any], json_path: Path) -> None:
    sheet_name = str(payload["sheet_name"])
    record = payload.get("record", {})
    if not isinstance(record, dict):
        record = {}
    headers = [str(item) for item in (payload.get("headers") or record.keys())]
    row = list(payload.get("row") or [record.get(header, "") for header in headers])
    record_id = str(payload.get("record_id") or json_path.stem)
    all_headers = headers + META_HEADERS
    all_row = _normalize_row(row, len(headers)) + [
        record_id,
        json_path.name,
        _now_text(),
    ]

    worksheet = _get_worksheet(workbook, sheet_name, all_headers)
    _dedupe_repeated_header_rows(worksheet, all_headers)
    _sync_headers(worksheet, all_headers)
    if _record_id_exists(worksheet, record_id):
        return
    worksheet.append(all_row)


def _get_worksheet(workbook: Workbook, sheet_name: str, headers: list[str]):
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        _ensure_header_row(worksheet, headers)
        return worksheet

    active = workbook.active
    if len(workbook.sheetnames) == 1 and _worksheet_empty(active):
        active.title = sheet_name
        _ensure_header_row(active, headers)
        return active

    worksheet = workbook.create_sheet(sheet_name)
    _ensure_header_row(worksheet, headers)
    return worksheet


def _worksheet_empty(worksheet: Any) -> bool:
    return worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet.cell(1, 1).value is None


def _ensure_header_row(worksheet: Any, headers: list[str]) -> None:
    if _worksheet_empty(worksheet):
        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=column_index, value=header)
        return
    _sync_headers(worksheet, headers)


def _sync_headers(worksheet: Any, headers: list[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        if worksheet.cell(row=1, column=column_index).value != header:
            worksheet.cell(row=1, column=column_index, value=header)


def _dedupe_repeated_header_rows(worksheet: Any, headers: list[str]) -> None:
    expected = [str(header) for header in headers]
    for row_index in range(worksheet.max_row, 1, -1):
        row_values = [
            str(worksheet.cell(row=row_index, column=column_index).value or "")
            for column_index in range(1, len(expected) + 1)
        ]
        if row_values == expected:
            worksheet.delete_rows(row_index, 1)


def _record_id_exists(worksheet: Any, record_id: str) -> bool:
    record_id_column = None
    for column_index in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=1, column=column_index).value == "记录ID":
            record_id_column = column_index
            break
    if record_id_column is None:
        return False
    for row_index in range(2, worksheet.max_row + 1):
        if str(worksheet.cell(row=row_index, column=record_id_column).value) == record_id:
            return True
    return False


def _save_workbook(workbook: Workbook, summary_paths: list[Path]) -> Path:
    errors: list[str] = []
    for path in summary_paths:
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            _format_workbook(workbook)
            workbook.save(path)
            return path
        except Exception as exc:
            errors.append(f"{path}: {exc}")
    raise RuntimeError("; ".join(errors) or "没有可用 summary 路径")


def _normalize_row(row: list[Any], length: int) -> list[Any]:
    normalized = list(row[:length])
    while len(normalized) < length:
        normalized.append("")
    return normalized


def _format_workbook(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        _format_worksheet(worksheet)


def _format_worksheet(worksheet: Any) -> None:
    if worksheet.max_row < 1:
        return

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, _display_width(str(value)))
        worksheet.column_dimensions[column_letter].width = min(
            MAX_COLUMN_WIDTH, max(MIN_COLUMN_WIDTH, max_length + 2)
        )


def _display_width(text: str) -> int:
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return width


@contextmanager
def _inbox_lock(inbox_dir: Path) -> Iterator[bool]:
    lock_path = inbox_dir / ".aggregate.lock"
    try:
        fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError:
        yield False
        return
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(_now_text())
        yield True
    finally:
        try:
            lock_path.unlink()
        except OSError:
            pass


def _app_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent.parent


def _now_text() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="汇总 Quick_Sparam 用户记录 JSON。")
    parser.add_argument(
        "--target",
        choices=["all", "local", "developer"],
        default="all",
        help="汇总目标：本地、开发者共享目录或全部。",
    )
    parser.add_argument(
        "--once",
        action="store_true",
        help="只汇总一次后退出。",
    )
    parser.add_argument(
        "--interval-seconds",
        type=int,
        default=0,
        help="大于 0 时按间隔循环汇总，适合常驻定时任务。",
    )
    args = parser.parse_args(argv)

    while True:
        stats = aggregate_records(target=args.target)
        print(
            f"本轮汇总完成：处理 {stats['processed']} 条，"
            f"删除 {stats['deleted']} 个 JSON，失败 {stats['failed']} 个。"
        )
        if args.once or args.interval_seconds <= 0:
            return 0
        time.sleep(args.interval_seconds)


if __name__ == "__main__":
    raise SystemExit(main())
