"""
后台 Excel 写入工具。

网络共享盘、UNC 路径或内网挂载点不可访问时，文件系统调用可能阻塞较久。
本模块把这些写入放到 daemon 后台线程里，避免影响主界面使用。
"""

from __future__ import annotations

import os
from pathlib import Path
import sys
from threading import Thread
from typing import Any

from openpyxl import Workbook, load_workbook


PATH_WRITE_TIMEOUT_SECONDS = 10


def append_excel_row_async(
    record_name: str,
    sheet_name: str,
    headers: list[str],
    row: list[Any],
    local_paths: list[Path],
    developer_paths: list[Path],
) -> None:
    print(f"{record_name}已提交，后台写入记录，不影响软件使用。")
    worker = Thread(
        target=_append_excel_row_worker,
        args=(record_name, sheet_name, headers, row, local_paths, developer_paths),
        daemon=True,
        name=f"{record_name}_excel_writer",
    )
    worker.start()


def configured_paths(mapping: dict[str, Any], app_base_dir: Path) -> list[Path]:
    configured = _platform_value(mapping)
    if configured is None:
        return []
    if isinstance(configured, (str, os.PathLike)):
        configured = [configured]

    paths: list[Path] = []
    for raw_path in configured:
        path = Path(os.path.expandvars(os.path.expanduser(str(raw_path))))
        if not path.is_absolute():
            path = app_base_dir / path
        paths.append(path)
    return paths


def _append_excel_row_worker(
    record_name: str,
    sheet_name: str,
    headers: list[str],
    row: list[Any],
    local_paths: list[Path],
    developer_paths: list[Path],
) -> None:
    _write_first_available_path(record_name, "本地", sheet_name, headers, row, local_paths)
    _write_first_available_path(
        record_name, "开发者", sheet_name, headers, row, developer_paths
    )


def _write_first_available_path(
    record_name: str,
    target_name: str,
    sheet_name: str,
    headers: list[str],
    row: list[Any],
    paths: list[Path],
) -> None:
    if not paths:
        print(f"提示：{record_name}{target_name}记录未配置写入路径。")
        return

    failures: list[tuple[Path, str]] = []
    for path in paths:
        ok, error = _append_excel_row_with_timeout(path, sheet_name, headers, row)
        if ok:
            print(f"{record_name}{target_name}记录已写入: {path}")
            return
        failures.append((path, error))

    print(
        f"提示：无法访问{record_name}{target_name}记录路径，已跳过该写入，"
        "不影响软件使用。"
    )
    for path, error in failures:
        print(f"  - {path}: {error}")


def _append_excel_row_with_timeout(
    path: Path, sheet_name: str, headers: list[str], row: list[Any]
) -> tuple[bool, str]:
    result: dict[str, Any] = {}

    def write_once() -> None:
        try:
            _append_excel_row(path, sheet_name, headers, row)
            result["ok"] = True
        except Exception as exc:
            result["error"] = str(exc)

    worker = Thread(target=write_once, daemon=True, name=f"excel_path_writer_{path.name}")
    worker.start()
    worker.join(PATH_WRITE_TIMEOUT_SECONDS)
    if worker.is_alive():
        return False, f"超过 {PATH_WRITE_TIMEOUT_SECONDS}s 未响应"
    if result.get("ok"):
        return True, ""
    return False, result.get("error", "未知写入错误")


def _append_excel_row(path: Path, sheet_name: str, headers: list[str], row: list[Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        workbook = load_workbook(path)
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(headers)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        worksheet.append(headers)

    if worksheet.max_row == 0 or worksheet.cell(row=1, column=1).value is None:
        worksheet.append(headers)
    else:
        _sync_headers(worksheet, headers)
    worksheet.append(row)
    workbook.save(path)


def _sync_headers(worksheet: Any, headers: list[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        if worksheet.cell(row=1, column=column_index).value != header:
            worksheet.cell(row=1, column=column_index, value=header)


def _platform_value(mapping: dict[str, Any]) -> Any:
    if sys.platform in mapping:
        return mapping[sys.platform]
    if sys.platform.startswith("linux") and "linux" in mapping:
        return mapping["linux"]
    return mapping.get("default")
