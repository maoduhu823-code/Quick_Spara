from datetime import datetime
import getpass
import subprocess
import sys
import os
import traceback

import numpy as np
import skrf as rf
from PyQt6.QtGui import QTextCursor
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QFileDialog, QListWidget, QLabel, QLineEdit, QMessageBox,
    QComboBox, QGroupBox, QCheckBox, QGridLayout,
    QTextEdit, QDialog, QInputDialog
)
from PyQt6.QtCore import Qt
from openpyxl import Workbook, load_workbook
import matplotlib
import matplotlib.pyplot as plt

from sparam_core import (enforce_nonzero_impedance, enforce_nonzero_z0,
                         SE2diff, SE2dq_dqs, SE2diff_port, parse_port_input,
                         merge_ports_multi, compute_time_domain)
from app_utils import show_error, check_and_set_port_names
from dialogs.cascade import CascadeDialog
from dialogs.freq_analysis import frequencyAnalysisDialog
from dialogs.se2diff import DiffConversionDialog
from dialogs.port_reduction import PortReductionDialog
from dialogs.port_reorder import PortOrderEditor
from dialogs.port_selector import PortSelector
from dialogs.port_name import PortNameDialog
from dialogs.loading import LoadingDialog
from dialogs.ripple import RippleFitDialog
from dialogs.port_merge import PortMergeDialog
from dialogs.port_management import PortManagementDialog, Z0EditDialog

# 参数类型 → 数据切面选项
_FACET_OPTIONS = {
    'S参数': ['幅度(dB)', '幅度(abs)', '相位(度)', '相位(rad)',
              'unwrap相位(度)', 'unwrap相位(rad)', '实部', '虚部', '群延迟(fs)'],
    'Y参数': ['导纳(abs)', '幅度(dB)', '相位(度)', '实部', '虚部'],
    'Z参数': ['阻抗(mΩ)', '幅度(dB)', '相位(度)', '实部(ESR)', '虚部', '电容(pF)'],
    '时域':  ['TDR阻抗', '阶跃响应', '冲激响应', '脉冲响应'],
}
# (参数类型, 数据切面) → (X轴缩放, Y轴缩放) 默认值
_DEFAULT_SCALES = {
    ('Z参数', '阻抗(mΩ)'): ('对数', '对数'),
    ('Y参数', '导纳(abs)'): ('对数', '对数'),
    ('Z参数', '实部(ESR)'): ('对数', '线性'),
}


class SParameterViewer_MainWin(QWidget):
    def __init__(self):
        super().__init__()
        self.start_time = datetime.now()
        self.user_name = getpass.getuser()
        self.initUI()
        self.plot_history = []
        self.current_plot_data = []
        self.s_data = {}
        self.s_param = {}
        self.y_param = {}
        self.z_param = {}
        self.loading = LoadingDialog(self)

    # ===== 数据缓存方法 =====

    def get_network(self, file_name):
        """从缓存获取或加载网络对象"""
        if file_name not in self.s_data:
            self.s_data[file_name] = rf.Network(file_name)
            enforce_nonzero_z0(self.s_data[file_name], file_name)
        return self.s_data[file_name]

    def get_s(self, file_name):
        """从缓存获取S参数矩阵"""
        if file_name not in self.s_param:
            try:
                self.s_param[file_name] = self.s_data[file_name].s
            except Exception as e:
                QMessageBox.warning(self, '加载错误', f"无法读取 {file_name}的S参数:\n{str(e)}")
                return None
        return self.s_param[file_name]

    def get_z(self, file_name):
        """从缓存获取Z参数矩阵"""
        if file_name not in self.z_param:
            try:
                self.z_param[file_name] = self.s_data[file_name].z
            except Exception as e:
                QMessageBox.warning(self, '加载错误', f"无法读取 {file_name}的Z参数:\n{str(e)}")
                return None
        return self.z_param[file_name]

    def get_y(self, file_name):
        """从缓存获取Y参数矩阵"""
        if file_name not in self.y_param:
            try:
                self.y_param[file_name] = self.s_data[file_name].y
            except Exception as e:
                QMessageBox.warning(self, '加载错误', f"无法读取 {file_name}的Y参数:\n{str(e)}")
                return None
        return self.y_param[file_name]

    def add_unique_filename(self, new_file_name):
        """向文件列表添加不重复的文件名，冲突时自动加数字后缀"""
        existing_names = {self.file_list.item(i).text() for i in range(self.file_list.count())}
        base_name = new_file_name
        suffix = 1
        while new_file_name in existing_names:
            if '.' in base_name:
                name_part, ext_part = base_name.rsplit('.', 1)
                new_file_name = f"{name_part}_{suffix}.{ext_part}"
            else:
                new_file_name = f"{base_name}_{suffix}"
            suffix += 1
        self.file_list.addItem(new_file_name)
        return new_file_name

    # ===== UI 构建 =====

    def initUI(self):
        self.version_num = 'B2026.1'
        self.setWindowTitle(f'Quick_Sparam_{self.version_num} 封装SIPI开发部'
                            ' --- 本工具免费提供给其他组织使用，但对出现的问题、结果等概不负责')
        self.setGeometry(100, 100, 1500, 700)

        main_layout = QHBoxLayout()
        main_layout.setSpacing(15)

        # ========== 左侧按钮栏 ==========
        left_panel = QVBoxLayout()
        left_panel.setContentsMargins(5, 5, 15, 5)
        left_panel.setSpacing(10)

        file_ops_group = QGroupBox("文件操作")
        file_ops_layout = QVBoxLayout()
        file_ops_layout.setSpacing(6)

        self.open_button = QPushButton('📂 打开S参数')
        self.open_button.setFixedHeight(38)
        self.open_button.clicked.connect(self.open_file_dialog)
        file_ops_layout.addWidget(self.open_button)

        self.save_button = QPushButton('💾 保存S参数')
        self.save_button.setFixedHeight(38)
        self.save_button.clicked.connect(self.save_sparameters)
        file_ops_layout.addWidget(self.save_button)

        self.read_button = QPushButton('🔍 查看源文件')
        self.read_button.setFixedHeight(38)
        self.read_button.clicked.connect(self.read_snp_file)
        file_ops_layout.addWidget(self.read_button)

        self.delete_button = QPushButton('🗑️ 删除S参数')
        self.delete_button.setFixedHeight(38)
        self.delete_button.clicked.connect(self.delete_selected_sparameters)
        file_ops_layout.addWidget(self.delete_button)

        file_ops_group.setLayout(file_ops_layout)
        left_panel.addWidget(file_ops_group)

        sparam_ops_group = QGroupBox("S参数操作")
        sparam_ops_layout = QVBoxLayout()
        sparam_ops_layout.setSpacing(4)

        self.port_management_button = QPushButton('端口处理')
        self.port_management_button.setFixedHeight(32)
        self.port_management_button.setToolTip(
            "端口元数据编辑（端口名/参考阻抗）、拓扑变换（重排/合并）、阻抗变换（重归一化）")
        self.port_management_button.clicked.connect(self.call_port_management)
        sparam_ops_layout.addWidget(self.port_management_button)

        self.cascade_button = QPushButton('S参数级联')
        self.cascade_button.setFixedHeight(32)
        self.cascade_button.clicked.connect(self.call_cascade)
        sparam_ops_layout.addWidget(self.cascade_button)

        self.diff_button = QPushButton('差分转换')
        self.diff_button.setFixedHeight(32)
        self.diff_button.clicked.connect(self.call_diff_conversion)
        sparam_ops_layout.addWidget(self.diff_button)

        self.analysis_btn = QPushButton('频域分析')
        self.analysis_btn.setFixedHeight(32)
        self.analysis_btn.clicked.connect(self.call_frequency_analysis_dialog)
        sparam_ops_layout.addWidget(self.analysis_btn)

        self.ripple_btn = QPushButton('Ripple拟合')
        self.ripple_btn.setFixedHeight(32)
        self.ripple_btn.clicked.connect(self.call_ripple_dialog)
        sparam_ops_layout.addWidget(self.ripple_btn)

        self.td_analysis_btn = QPushButton('时域分析')
        self.td_analysis_btn.setFixedHeight(32)
        self.td_analysis_btn.clicked.connect(self.call_time_domain_dialog)
        sparam_ops_layout.addWidget(self.td_analysis_btn)

        sparam_ops_group.setLayout(sparam_ops_layout)
        left_panel.addWidget(sparam_ops_group)

        left_panel.addStretch()

        # ========== 右侧内容区 ==========
        right_panel = QVBoxLayout()
        right_panel.setContentsMargins(5, 5, 5, 5)

        file_group = QGroupBox("S参数文件列表")
        file_layout = QVBoxLayout()
        self.file_list = QListWidget()
        self.file_list.setSelectionMode(QListWidget.SelectionMode.ExtendedSelection)
        self.file_list.setMinimumWidth(400)
        file_layout.addWidget(self.file_list)
        file_group.setLayout(file_layout)
        right_panel.addWidget(file_group, stretch=4)

        # 绘图控制
        plot_group = QGroupBox("绘图控制")
        plot_layout = QHBoxLayout()

        plot_left_layout = QGridLayout()
        plot_left_layout.addWidget(QLabel("端口1:"), 0, 0)
        self.port1_input = QLineEdit()
        self.port1_input.setPlaceholderText("例: 1 2 3 或 1:5")
        plot_left_layout.addWidget(self.port1_input, 0, 1, 1, 2)

        plot_left_layout.addWidget(QLabel("端口2:"), 1, 0)
        self.port2_input = QLineEdit()
        self.port2_input.setPlaceholderText("例: 1:2:5")
        plot_left_layout.addWidget(self.port2_input, 1, 1, 1, 2)

        self.plot_button = QPushButton('绘 图')
        self.plot_button.setFixedHeight(50)
        self.plot_button.clicked.connect(self.plot_s_parameters)
        plot_left_layout.addWidget(self.plot_button, 3, 0, 2, 3)

        plot_right_layout = QGridLayout()
        self.port_select_btn = QPushButton("选择端口")
        self.port_select_btn.clicked.connect(self.on_port_select)
        self.port_select_btn.setFixedWidth(80)
        plot_right_layout.addWidget(self.port_select_btn, 0, 0)

        self.btn_freq_list = QPushButton('频点列表')
        self.btn_freq_list.clicked.connect(self.print_freq_axis)
        plot_right_layout.addWidget(self.btn_freq_list, 0, 1)

        self.Basic_info = QPushButton('文件信息')
        self.Basic_info.clicked.connect(self.Basic_info_print)
        plot_right_layout.addWidget(self.Basic_info, 0, 2)

        self.BandWidth_focus = QPushButton('频域切片')
        self.BandWidth_focus.clicked.connect(self.call_freq_slice)
        plot_right_layout.addWidget(self.BandWidth_focus, 0, 3)

        plot_right_layout.addWidget(QLabel("映射模式:"), 1, 0)
        self.mapping_combo = QComboBox()
        self.mapping_combo.addItems(["一 一对应", "交叉映射"])
        plot_right_layout.addWidget(self.mapping_combo, 1, 1, 1, 3)

        plot_right_layout.addWidget(QLabel("参数类型:"), 2, 0)
        self.param_type_combo = QComboBox()
        self.param_type_combo.addItems(['S参数', 'Y参数', 'Z参数', '时域'])
        plot_right_layout.addWidget(self.param_type_combo, 2, 1, 1, 3)

        plot_right_layout.addWidget(QLabel("数据切面:"), 3, 0)
        self.facet_combo = QComboBox()
        plot_right_layout.addWidget(self.facet_combo, 3, 1, 1, 3)

        self._xlbl = QLabel("X轴:")
        plot_right_layout.addWidget(self._xlbl, 4, 0)
        self.xscale_combo = QComboBox()
        self.xscale_combo.addItems(['线性', '对数'])
        self.xscale_combo.setFixedWidth(60)
        plot_right_layout.addWidget(self.xscale_combo, 4, 1)
        self._ylbl = QLabel("Y轴:")
        plot_right_layout.addWidget(self._ylbl, 4, 2)
        self.yscale_combo = QComboBox()
        self.yscale_combo.addItems(['线性', '对数'])
        self.yscale_combo.setFixedWidth(60)
        plot_right_layout.addWidget(self.yscale_combo, 4, 3)

        # 时域专属参数（默认隐藏，选"时域"时显示）
        self._td_tr_lbl = QLabel("t_rise (ps):")
        self._td_tr_edit = QLineEdit("50")
        self._td_tr_edit.setFixedWidth(65)
        self._td_dt_lbl = QLabel("t_step (ps):")
        self._td_dt_edit = QLineEdit("25")
        self._td_dt_edit.setFixedWidth(65)
        plot_right_layout.addWidget(self._td_tr_lbl,  6, 0)
        plot_right_layout.addWidget(self._td_tr_edit, 6, 1)
        plot_right_layout.addWidget(self._td_dt_lbl,  6, 2)
        plot_right_layout.addWidget(self._td_dt_edit, 6, 3)

        self._td_z0_lbl = QLabel("Z0 (Ω):")
        self._td_z0_edit = QLineEdit("50")
        self._td_z0_edit.setFixedWidth(65)
        plot_right_layout.addWidget(self._td_z0_lbl,  7, 0)
        plot_right_layout.addWidget(self._td_z0_edit, 7, 1)

        self._xy_widgets = [self._xlbl, self.xscale_combo, self._ylbl, self.yscale_combo]
        self._td_widgets = [self._td_tr_lbl, self._td_tr_edit,
                            self._td_dt_lbl, self._td_dt_edit,
                            self._td_z0_lbl, self._td_z0_edit]
        for w in self._td_widgets:
            w.setVisible(False)

        self.legend_checkbox = QCheckBox("显示图例")
        self.legend_checkbox.setChecked(True)
        plot_right_layout.addWidget(self.legend_checkbox, 5, 0)

        self.same_plot_checkbox = QCheckBox("曲线叠加")
        plot_right_layout.addWidget(self.same_plot_checkbox, 5, 1)

        plot_right_layout.addWidget(QLabel("关注频点:"), 5, 2)
        self.freG_input = QLineEdit("")
        self.freG_input.setPlaceholderText("GHz")
        plot_right_layout.addWidget(self.freG_input, 5, 3)

        # 初始化切面选项并连接信号
        self._update_facet_options()
        self.param_type_combo.currentIndexChanged.connect(self._update_facet_options)
        self.facet_combo.currentIndexChanged.connect(self._update_default_scales)

        plot_layout.addLayout(plot_left_layout, stretch=2)
        plot_layout.addLayout(plot_right_layout, stretch=2)
        plot_group.setLayout(plot_layout)

        right_panel.addWidget(plot_group, stretch=2)

        main_layout.addLayout(left_panel, stretch=1)
        main_layout.addLayout(right_panel, stretch=4)

        # 信息输出栏
        down_panel = QVBoxLayout()
        output_group = QGroupBox("信息输出")
        output_layout = QVBoxLayout()
        output_text_layout = QHBoxLayout()
        self.output_console = QTextEdit()
        self.output_console.setReadOnly(True)
        output_text_layout.addWidget(self.output_console)

        output_button_layout = QHBoxLayout()
        clear_cache_button = QPushButton("清除缓存")
        clear_cache_button.clicked.connect(self.clear_cache_data)
        output_button_layout.addWidget(clear_cache_button)
        clear_button = QPushButton("清除输出")
        clear_button.clicked.connect(self.output_console.clear)
        output_button_layout.addWidget(clear_button)
        save_button = QPushButton("保存输出")
        save_button.clicked.connect(self.save_output_to_file)
        output_button_layout.addWidget(save_button)
        info_button = QPushButton("版本信息")
        info_button.clicked.connect(self.info_version)
        output_button_layout.addWidget(info_button)

        output_layout.addLayout(output_text_layout)
        output_layout.addLayout(output_button_layout)
        output_group.setLayout(output_layout)
        down_panel.addWidget(output_group)

        self._original_stdout = sys.stdout
        sys.stdout = self
        main_layout.addLayout(down_panel, stretch=2)

        self.setLayout(main_layout)
        self._setup_ui_style()

    # ===== UI 支撑 =====

    def _setup_ui_style(self):
        button_style = """
        QPushButton {
            font-size: 14px;
            padding: 8px;
            border-radius: 5px;
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                      stop:0 #f6f7fa, stop:1 #dadbde);
            min-width: 80px;
        }
        QPushButton:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                      stop:0 #e7e8eb, stop:1 #cbcccf);
        }
        """
        for btn in [self.open_button, self.save_button, self.diff_button,
                    self.port_management_button, self.cascade_button,
                    self.delete_button, self.analysis_btn, self.plot_button,
                    self.ripple_btn, self.td_analysis_btn, self.read_button]:
            btn.setStyleSheet(button_style)

        self.file_list.setStyleSheet("""
            QListWidget {
                font-size: 13px;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
            QListWidget::item {
                padding: 5px;
            }
        """)

    def write(self, text):
        cursor = self.output_console.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        cursor.insertText(text)
        self.output_console.setTextCursor(cursor)
        self.output_console.ensureCursorVisible()

    def check_beta_period(self):
        beta_end_date = datetime(2026, 6, 30)
        if datetime.now() > beta_end_date:
            QMessageBox.critical(self, "内测结束",
                                 "内测已结束，请联系管理员获取公测版本。",
                                 QMessageBox.StandardButton.Ok)
            return False
        return True

    def closeEvent(self, event):
        self.on_app_closing()
        super().closeEvent(event)

    def on_app_closing(self):
        close_time = datetime.now()
        usage_duration = str(close_time - self.start_time)
        data_to_write = [self.user_name, self.start_time.strftime("%Y-%m-%d"), usage_duration]
        print("写入使用日志:", data_to_write)
        if sys.platform != 'win32':
            self.write_usage_to_network_excel(
                shared_folder=r"/data/Storage_pisi/w00810255/Qs",
                filename="Quick_Sparam_usage_log.xlsx",
                data=data_to_write
            )

    def write_usage_to_network_excel(self, shared_folder, filename, data):
        unc_path = os.path.join(shared_folder, filename)
        try:
            if os.path.exists(unc_path):
                wb = load_workbook(unc_path)
                ws = wb.active
                if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
                    ws.append(["用户名", "日期", "使用时长"])
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(["用户名", "日期", "使用时长"])
            ws.append(data)
            wb.save(unc_path)
            print(f"✅ 成功写入: {unc_path}")
        except PermissionError:
            print("❌ 无法写入文件，文件可能正在被其他用户打开。")
        except Exception as e:
            print(f"❌ 写入Excel失败: {e}")

    # ===== 主界面功能 =====

    def clear_cache_data(self):
        self.file_list.clear()
        self.s_data.clear()

    def save_output_to_file(self):
        output_text = self.output_console.toPlainText()
        if not output_text.strip():
            QMessageBox.warning(self, '无内容', '输出框中没有内容可保存！')
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存输出内容", "", "文本文件 (*.txt);;所有文件 (*)"
        )
        if file_path:
            try:
                if not file_path.lower().endswith('.txt'):
                    file_path += '.txt'
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(output_text)
                QMessageBox.information(self, '保存成功', f'输出内容已保存到:\n{file_path}')
            except Exception:
                show_error(self, "保存文件时出错")

    def info_version(self):
        print("B2026版主要更新内容如下：")
        items = [
            "---功能增加",
            "【频域分析】   功能中加入指定bit、最差bit的频域结果查看和比对",
            "【频域分析】   新增了VTF loss/crosstalk的数据类型",
            "【端口缩并】   新增了Cio端接的功能",
            "---交互体验",
            "【文件列表】   改为更为常用的Ctrl/Shift+点击的交互模式",
            "【端口重排】   支持端口批量选中、拖拽实现顺序调整",
            "【频域分析】   修复保存excel时文件名的长度限制",
            "【参数级联】   修复了特定级联后S参数保存的bug",
            "【参数读取】   HFSS生成的频变阻抗信息的S参数会导致报错，端口阻抗统一修复为标量阻抗",
        ]
        for text in items:
            print(f"{text}")
        print(f'\n版本持续迭代中，当前版本号为{self.version_num}')

    def open_file_dialog(self):
        file_names, _ = QFileDialog.getOpenFileNames(
            self, '打开 S 参数文件', '', 'S 参数文件 (*.s*p *.S*P)'
        )
        if file_names:
            for file_name in file_names:
                self.file_list.addItem(file_name)

    def read_snp_file(self):
        selected_files = [item.text() for item in self.file_list.selectedItems()]
        for file in selected_files:
            if sys.platform == 'win32':
                subprocess.Popen(['notepad.exe', file])
            else:
                subprocess.Popen(['gvim', file])

    def save_sparameters(self):
        selected_items = self.file_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, '警告', '请先在列表中选择要保存的S参数文件！')
            return
        save_dir = QFileDialog.getExistingDirectory(
            self, "选择保存目录", "", QFileDialog.Option.ShowDirsOnly
        )
        if not save_dir:
            return
        success_count = 0
        for item in selected_items:
            try:
                file_path = item.text()
                network = self.get_network(file_path)
                file_name = os.path.basename(file_path)
                save_path = os.path.join(save_dir, file_name)
                if os.path.exists(save_path):
                    base, ext = os.path.splitext(file_name)
                    counter = 1
                    while os.path.exists(os.path.join(save_dir, f"{base}_{counter}{ext}")):
                        counter += 1
                    save_path = os.path.join(save_dir, f"{base}_{counter}{ext}")

                z0_array = np.atleast_1d(network.z0)
                all_equal_z0 = np.allclose(z0_array, z0_array[0, 0])
                if all_equal_z0:
                    try:
                        print('尝试使用默认方法保存（touchstone1.0）')
                        network.write_touchstone(save_path)
                    except UnicodeEncodeError:
                        print('默认方法失败，尝试使用自定义方法保存（touchstone2.0）')
                        ts_string = network.write_touchstone(return_string=True, write_z0=True)
                        with open(save_path, 'w', encoding='utf-8') as f:
                            f.write(ts_string)
                else:
                    print("阻抗不一致，尝试使用自定义方法保存（touchstone2.0）")
                    ts_string = network.write_touchstone(return_string=True, write_z0=True)
                    with open(save_path, 'w', encoding='utf-8') as f:
                        f.write(ts_string)
                success_count += 1
            except Exception as e:
                QMessageBox.warning(self, '保存失败',
                                    f'文件 {os.path.basename(file_path)} 保存失败: {str(e)}')
        if success_count > 0:
            QMessageBox.information(self, '保存成功',
                                    f'成功保存 {success_count}/{len(selected_items)} 个文件到:\n{save_dir}')

    def delete_selected_sparameters(self):
        selected_items = self.file_list.selectedItems()
        if not selected_items:
            return
        for item in selected_items:
            file_name = item.text()
            self.file_list.takeItem(self.file_list.row(item))
            if file_name in self.s_data:
                del self.s_data[file_name]

    def on_curve_click(self, event):
        if event.artist not in self.plot_lines:
            print('No line data')
            return
        line = event.artist
        self.highlight_curve(line)
        self.print_curve_info(line)
        self.fig.canvas.draw()

    def highlight_curve(self, selected_line):
        for line in self.plot_lines:
            line.set_linewidth(1 if line != selected_line else 3)
            line.set_alpha(0.7 if line != selected_line else 1.0)
        legend = self.ax.get_legend()
        if legend:
            for text in legend.get_texts():
                if text.get_text() == selected_line.get_label():
                    text.set_fontweight('bold')
                    text.set_color('red')
                else:
                    text.set_fontweight('normal')
                    text.set_color('black')

    def print_curve_info(self, line):
        info = f"""
        === 曲线信息 ===
        文件名: {line.network_name}
        参数: S{line.port_pair[0]},{line.port_pair[1]}
        模式: {line.data_mode}
        """
        print(info)

    def _update_facet_options(self):
        param_type = self.param_type_combo.currentText()
        facets = _FACET_OPTIONS.get(param_type, [])
        self.facet_combo.blockSignals(True)
        self.facet_combo.clear()
        self.facet_combo.addItems(facets)
        self.facet_combo.blockSignals(False)
        is_td = (param_type == '时域')
        for w in self._xy_widgets:
            w.setVisible(not is_td)
        for w in self._td_widgets:
            w.setVisible(is_td)
        if is_td:
            self._refresh_td_defaults()
        self._update_default_scales()

    def _refresh_td_defaults(self):
        """根据当前选中文件自动设置 t_rise / t_step 的合理默认值。"""
        from sparam_core import td_default_params
        selected = [item.text() for item in self.file_list.selectedItems()]
        if not selected:
            selected = [self.file_list.item(i).text()
                        for i in range(self.file_list.count())]
        networks = [self.s_data[f] for f in selected if f in self.s_data]
        if not networks:
            return
        try:
            defs = [td_default_params(n) for n in networks]
            tr = max(d['tr_ps'] for d in defs)
            dt = max(d['dt_ps'] for d in defs)
            self._td_tr_edit.setText(f'{tr:.2f}')
            self._td_dt_edit.setText(f'{dt:.2f}')
        except Exception:
            pass

    def _update_default_scales(self):
        param_type = self.param_type_combo.currentText()
        facet = self.facet_combo.currentText()
        if param_type == '时域':
            # Z0 仅对 TDR 有意义，其余模式置灰
            is_tdr = (facet == 'TDR阻抗')
            self._td_z0_lbl.setEnabled(is_tdr)
            self._td_z0_edit.setEnabled(is_tdr)
            return
        xscale, yscale = _DEFAULT_SCALES.get((param_type, facet), ('线性', '线性'))
        self.xscale_combo.setCurrentText(xscale)
        self.yscale_combo.setCurrentText(yscale)

    def get_current_plot_config(self):
        return (self.param_type_combo.currentText(),
                self.facet_combo.currentText(),
                self.xscale_combo.currentText(),
                self.yscale_combo.currentText())

    def _plot_single_curve(self, network, szy_params, p1, p2):
        num_port = network.number_of_ports
        if p1 > num_port or p2 > num_port:
            QMessageBox.warning(self, '端口错误',
                                f'文件 {network.name} 的端口{p1}或{p2}超出范围！')
            return
        param_type, facet, xscale, yscale = self.get_current_plot_config()
        freqG = network.frequency.f / 1e9
        label = f'{network.name}_S{p1},{p2}'

        # 时域：使用时间轴，提前返回
        if param_type == '时域':
            td_mode_map = {'TDR阻抗': 'TDR', '阶跃响应': 'step',
                           '冲激响应': 'impulse', '脉冲响应': 'pulse'}
            try:
                _tr = float(self._td_tr_edit.text())
            except ValueError:
                _tr = None
            try:
                _dt = float(self._td_dt_edit.text())
            except ValueError:
                _dt = None
            try:
                _z0 = float(self._td_z0_edit.text())
            except ValueError:
                _z0 = 50.0
            result = compute_time_domain(
                network, p1, p2, td_mode_map.get(facet, 'TDR'),
                tr_ps=_tr, dt_ps=_dt, z0=_z0
            )
            x_td = result["time_ps"]
            y_td = result["y_data"]
            lbl = f'{network.name}_{facet}_S{p1},{p2}'
            if self.legend_checkbox.isChecked():
                line, = self.ax.plot(x_td, y_td, label=lbl, picker=5)
            else:
                line, = self.ax.plot(x_td, y_td, picker=5)
            line.network_name = network.name
            line.port_pair = (p1, p2)
            line.data_mode = f'{param_type} {facet}'
            line.freq_data = x_td
            line.value_data = y_td
            return line

        param = szy_params[:, p1 - 1, p2 - 1]

        if facet == '幅度(dB)':
            y_data = 20 * np.log10(np.abs(param))
        elif facet in ('幅度(abs)', '导纳(abs)'):
            y_data = np.abs(param)
        elif facet == '阻抗(mΩ)':
            y_data = 1000 * np.abs(param)
        elif facet == '相位(度)':
            y_data = np.angle(param) * 180 / np.pi
        elif facet == '相位(rad)':
            y_data = np.angle(param)
        elif facet == 'unwrap相位(度)':
            y_data = np.unwrap(np.angle(param)) * 180 / np.pi
        elif facet == 'unwrap相位(rad)':
            y_data = np.unwrap(np.angle(param))
        elif facet == '群延迟(fs)':
            phase = np.unwrap(np.angle(param))
            tau_g = -np.gradient(phase, freqG * 1e9) / (2 * np.pi)
            y_data = tau_g * 1e15
        elif facet in ('实部', '实部(ESR)'):
            y_data = np.real(param)
        elif facet == '虚部':
            y_data = np.imag(param)
        elif facet == '电容(pF)':
            with np.errstate(divide='ignore', invalid='ignore'):
                y_data = -1.0 / (2 * np.pi * freqG * 1e9 * np.imag(param)) * 1e12
            label = f'{network.name}_C{p1},{p2}'
        else:
            y_data = np.abs(param)

        if self.legend_checkbox.isChecked():
            line, = self.ax.plot(freqG, y_data, label=label, picker=5)
        else:
            line, = self.ax.plot(freqG, y_data, picker=5)

        line.network_name = network.name
        line.port_pair = (p1, p2)
        line.data_mode = f'{param_type} {facet}'
        line.freq_data = freqG
        line.value_data = y_data

        self.ax.set_xscale('log' if xscale == '对数' else 'linear')
        self.ax.set_yscale('log' if yscale == '对数' else 'linear')

        input_text = self.freG_input.text().split()
        mark_freqGs = list(map(float, input_text))
        if mark_freqGs:
            for fG_mark in mark_freqGs:
                if min(freqG) <= fG_mark <= max(freqG):
                    idx = np.abs(freqG - fG_mark).argmin()
                    actual_freq = freqG[idx]
                    actual_value = y_data[idx]
                    self.ax.axvline(x=actual_freq, linestyle=':', alpha=1)
                    mask = (freqG <= fG_mark)
                    band_freqG = freqG[mask]
                    band_data = y_data[mask]
                    max_idx = np.argmax(band_data)
                    min_idx = np.argmin(band_data)
                    self.ax.plot(band_freqG[max_idx], band_data[max_idx], 'ro', markersize=5)
                    self.ax.plot(band_freqG[min_idx], band_data[min_idx], 'ro', markersize=5)
                    print(f"S{p1},{p2} {param_type} {facet} 信息提取---\n"
                          f"频率范围: {band_freqG[0]:.1f}-{band_freqG[-1]:.1f} GHz\n"
                          f"频率——max: {band_freqG[max_idx]:.3f}, min: {band_freqG[min_idx]:.3f}, end: {actual_freq:.3f}\n"
                          f"数值——max: {band_data[max_idx]:.3f}; min: {band_data[min_idx]:.3f}, end: {actual_value:.3f}\n")
                else:
                    print('关注频点超出现有的频率数据，无法进行数据统计')
        return line

    def plot_s_parameters(self):
        selected_files = [item.text() for item in self.file_list.selectedItems()]
        port1 = self.port1_input.text().strip()
        port2 = self.port2_input.text().strip()
        mapping_mode = self.mapping_combo.currentText()

        if not selected_files or not port1 or not port2:
            QMessageBox.warning(self, '输入错误', '请在上方列表中选择文件! 并输入端口1和端口2！')
            return

        port1_list = parse_port_input(port1)
        port2_list = parse_port_input(port2)
        if port1_list is None or port2_list is None:
            return

        new_plot_data = {'files': selected_files, 'port1': port1_list,
                         'port2': port2_list, 'mode': mapping_mode}
        if not self.same_plot_checkbox.isChecked():
            self.plot_history = [new_plot_data]
        else:
            self.plot_history.append(new_plot_data)

        self.fig, self.ax = plt.subplots()
        if sys.platform == 'win32':
            plt.rcParams['font.sans-serif'] = ['SimHei']
        else:
            plt.rcParams['font.sans-serif'] = ['WenQuanYi Zen Hei']
        plt.rcParams['axes.unicode_minus'] = False
        param_type, facet, _, _ = self.get_current_plot_config()
        if param_type == '时域':
            _td_ylabels = {'TDR阻抗': '阻抗 (Ω)', '阶跃响应': '阶跃响应',
                           '冲激响应': 'h(t)', '脉冲响应': '脉冲响应'}
            self.ax.set_xlabel("时间 (ps)")
            self.ax.set_ylabel(_td_ylabels.get(facet, '时域'))
        else:
            self.ax.set_xlabel("频率 (GHz)")
            self.ax.set_ylabel(f"{param_type} {facet}")
        self.ax.grid(True)

        self.loading = LoadingDialog(self)
        self.loading.show()
        QApplication.processEvents()

        try:
            self.plot_lines = []
            for plot_data in self.plot_history:
                for file_name in plot_data['files']:
                    if self.loading.cancelled:
                        break
                    show_str = os.path.basename(file_name)
                    self.loading.set_message(f"正在处理文件: {show_str}")
                    QApplication.processEvents()
                    try:
                        network = self.get_network(file_name)
                        port1_plot = plot_data['port1']
                        port2_plot = plot_data['port2']
                        print(f"{file_name}")

                        def _get_params(fname):
                            if param_type == 'Z参数':
                                return self.get_z(fname)
                            elif param_type == 'Y参数':
                                return self.get_y(fname)
                            else:
                                return self.get_s(fname)

                        if mapping_mode == "一 一对应":
                            if len(port1_plot) != len(port2_plot):
                                QMessageBox.warning(self, '输入错误', '一一对应模式需要端口数量相同！')
                                return
                            for p1, p2 in zip(port1_plot, port2_plot):
                                line = self._plot_single_curve(network, _get_params(file_name), p1, p2)
                                self.plot_lines.append(line)
                        elif mapping_mode == "交叉映射":
                            for p1 in port1_plot:
                                for p2 in port2_plot:
                                    line = self._plot_single_curve(network, _get_params(file_name), p1, p2)
                                    self.plot_lines.append(line)
                    except Exception as e:
                        show_error(self, f"处理文件 {file_name} 时出错: {str(e)}")
                        continue

            if facet == '电容(pF)':
                self.ax.text(
                    0.01, 0.01,
                    r"$C = \frac{-1}{2\pi f \cdot \mathrm{Im}[Z_{ij}]}$  (pF)",
                    transform=self.ax.transAxes,
                    fontsize=10, color='gray', va='bottom', ha='left',
                    bbox=dict(boxstyle='round,pad=0.2', fc='white', alpha=0.6, ec='none')
                )
            self.ax.legend()
            if hasattr(self, '_cid'):
                self.fig.canvas.mpl_disconnect(self._cid)
            self._cid = self.fig.canvas.mpl_connect('pick_event', self.on_curve_click)
            self.fig.canvas.draw()
            self.fig.show()
        except Exception as e:
            show_error(self, f"绘图时遇到错误: {str(e)}")
        finally:
            if hasattr(self, 'loading'):
                self.loading.close()

    def on_port_select(self):
        try:
            target_input = self.port2_input if self.port2_input.hasFocus() else self.port1_input
            file_list = [a.text() for a in self.file_list.selectedItems()]
            selected_ports = check_and_set_port_names(self, file_list)
            if selected_ports:
                target_input.setText(" ".join(map(str, selected_ports)))
        except Exception as e:
            show_error(self, f"端口选择出错: {str(e)}")

    def call_freq_slice(self):
        selected_files = [item.text() for item in self.file_list.selectedItems()]
        port1 = self.port1_input.text().strip()
        port2 = self.port2_input.text().strip()
        if not selected_files or not port1 or not port2:
            QMessageBox.warning(self, '输入错误', '请选择文件并在端口1/2栏输入端口！')
            return
        port1_list = parse_port_input(port1)
        port2_list = parse_port_input(port2)
        if port1_list is None or port2_list is None:
            return

        text, ok = QInputDialog.getText(self, '频域切片', '输入频率范围 (GHz)，如：1~5')
        if not ok or not text.strip():
            return
        try:
            parts = text.strip().replace('~', ' ').split()
            f_min, f_max = float(parts[0]), float(parts[1])
        except Exception:
            QMessageBox.warning(self, '格式错误', '请输入如 "1~5" 的格式')
            return

        if sys.platform == 'win32':
            plt.rcParams['font.sans-serif'] = ['SimHei']
        else:
            plt.rcParams['font.sans-serif'] = ['WenQuanYi Zen Hei']
        plt.rcParams['axes.unicode_minus'] = False

        for file_name in selected_files:
            try:
                network = self.get_network(file_name)
                s = self.get_s(file_name)
                freqG = network.f / 1e9
                mask = (freqG >= f_min) & (freqG <= f_max)
                freqG_sl = freqG[mask]
                if freqG_sl.size == 0:
                    print(f'{file_name}: 所选频率范围内无频点')
                    continue

                fig, axes = plt.subplots(2, 2, figsize=(10, 7))
                fig.suptitle(f'{os.path.basename(file_name)}  [{f_min}~{f_max} GHz]')

                pairs = (list(zip(port1_list, port2_list))
                         if len(port1_list) == len(port2_list)
                         else [(p1, p2) for p1 in port1_list for p2 in port2_list])

                for p1, p2 in pairs:
                    if p1 > network.nports or p2 > network.nports:
                        continue
                    param_sl = s[mask, p1 - 1, p2 - 1]
                    lbl = f'S{p1},{p2}'
                    axes[0, 0].plot(freqG_sl, 20 * np.log10(np.abs(param_sl)), label=lbl)
                    axes[0, 1].plot(freqG_sl, np.angle(param_sl) * 180 / np.pi, label=lbl)
                    axes[1, 0].plot(freqG_sl, np.real(param_sl), label=lbl)
                    axes[1, 1].plot(freqG_sl, np.imag(param_sl), label=lbl)

                titles = ['幅度 (dB)', '相位 (度)', '实部', '虚部']
                ylabels = ['dB', '度', '', '']
                for ax, title, ylabel in zip(axes.flat, titles, ylabels):
                    ax.set_title(title)
                    ax.set_ylabel(ylabel)
                    ax.set_xlabel('频率 (GHz)')
                    ax.grid(True)
                    ax.legend()
                plt.tight_layout()
                plt.show()
            except Exception as e:
                show_error(self, f"频域切片出错 ({file_name}): {str(e)}")

    def print_freq_axis(self):
        selected_items = self.file_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, '错误', '请先选择文件！')
            return
        for item in selected_items:
            try:
                network = self.get_network(item.text())
                freqs = network.f
                print(f"\n=== {os.path.basename(item.text())} 频率轴 ===")
                print(f"{'序号':>5}  {'频率(GHz)':>14}  {'间距(MHz)':>10}")
                print("-" * 38)
                for i, f in enumerate(freqs):
                    spacing = (freqs[i] - freqs[i - 1]) / 1e6 if i > 0 else 0.0
                    print(f"{i + 1:>5}  {f / 1e9:>14.6f}  {spacing:>10.3f}")
                print(f"共 {len(freqs)} 个频点，范围 {freqs[0] / 1e9:.4f} ~ {freqs[-1] / 1e9:.4f} GHz")
            except Exception as e:
                show_error(self, f"打印频率轴出错: {str(e)}")

    def Basic_info_print(self):
        try:
            selected_items = self.file_list.selectedItems()
            for item in selected_items:
                file_name = item.text()
                S = self.get_network(file_name)
                z0 = S.z0[0, :]
                freq = S.f / 1e9
                num_port = S.number_of_ports
                portnames = S.port_names
                print(f'文件名：{file_name}')
                print(f'频率范围：{freq[0]} ~ {freq[-1]} GHz')
                print(f'频点数：{len(freq)}')
                print(f'端口数量：{num_port}')
                if portnames is not None and len(portnames) > 0:
                    print('端口名称：')
                    for i, (port_name, zref) in enumerate(zip(portnames, abs(z0)), start=1):
                        print(f'Port {i}: {port_name},   Z_ref={int(zref)}')
                else:
                    print(f'该S参数没有可用的端口名称信息')
        except Exception:
            show_error(self, "打印信息时出错")

    # ===== 次级界面功能 =====

    def call_diff_conversion(self):
        selected_files = [item.text() for item in self.file_list.selectedItems()]
        if not selected_files:
            QMessageBox.warning(self, '错误', '请先选择文件！')
            return

        network = self.get_network(selected_files[0])
        if not network:
            return
        dialog = DiffConversionDialog(network.nports)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            try:
                params = dialog.get_conversion_params()
                if not params:
                    return
                print("转换参数:")
                [print(f"{k} = {v}") for k, v in params.items()]

                z0_diff = params['z0_diff']
                port_mode = params['port_mode']
                output_mode = params['output_mode']
                logic_mode = params['Logic']

                for file_name in selected_files:
                    try:
                        network = self.get_network(file_name)
                        if not network:
                            continue
                        if logic_mode == 'line':
                            if 'diff_list' in params and params['diff_list']:
                                new_network = SE2dq_dqs(network, params['diff_list'], port_mode, z0_diff)
                            else:
                                new_network = SE2diff(network, port_mode, output_mode, z0_diff)
                        else:
                            new_network = SE2diff_port(network, params['diff_list'],
                                                       params['z0_diff'], params['output_mode'])
                        if new_network:
                            new_name = self.add_unique_filename(new_network.name)
                            self.s_data[new_name] = new_network
                            if new_network.port_names:
                                print('转换后的端口名称:')
                                print(*new_network.port_names, sep="\n")
                    except Exception as e:
                        QMessageBox.critical(self, "转换错误",
                                             f"处理文件 {file_name} 时出错:\n{str(e)}")
                        traceback.print_exc()
            except Exception as e:
                QMessageBox.critical(self, "参数错误", f"获取参数时出错:\n{str(e)}")
                traceback.print_exc()

    def call_port_reduction(self):
        try:
            selected_files = [item.text() for item in self.file_list.selectedItems()]
            if not selected_files:
                QMessageBox.warning(self, '错误', '请先选择文件！')
                return
            if len(selected_files) > 1:
                QMessageBox.information(self, "提示", "检测到多文件选择，请确定选择的S参数以同样的配置处理端口")

            dialog = PortReductionDialog(self, selected_files)
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return

            port_configs, z0_configs, port_delete = dialog.get_result()
            print('端口设置详情')
            print(f'端口组：{port_configs}')
            print(f'端接配置(R,C)：{z0_configs}')
            print(f'简并端口：{port_delete}')

            for file_name in selected_files:
                try:
                    network_ori = self.get_network(file_name)
                    enforce_nonzero_impedance(network_ori)
                    network = network_ori.copy()
                    n_ports = network.nports
                    freq = network.frequency.f
                    file_s_name = network.name

                    all_ports = [p for group in port_configs for p in group]
                    if all_ports and max(all_ports) > n_ports:
                        QMessageBox.warning(self, '错误',
                                            f'文件 {file_s_name} 的端口范围超出！最大端口数为 {n_ports}')
                        continue
                    if port_delete and max(port_delete) > n_ports:
                        QMessageBox.warning(self, '错误',
                                            f'文件 {file_s_name} 要删除的端口超出范围！最大端口数为 {n_ports}')
                        continue

                    z0_new = network.z0.copy()
                    for ports, z_rc in zip(port_configs, z0_configs):
                        port_indices = [p - 1 for p in ports]
                        Z_load = dialog.compute_load_impedance(freq, z_rc)
                        for p in port_indices:
                            z0_new[:, p] = Z_load

                    network.renormalize(z0_new)

                    if port_delete:
                        keep_ports = [i for i in range(n_ports) if (i + 1) not in port_delete]
                        network = network.subnetwork(ports=keep_ports)

                    all_c_zero = all(c == 0 for (_, c) in z0_configs)
                    suffix = f'_renorm_R.s{network.nports}p' if all_c_zero else f'_renorm_RC.s{network.nports}p'
                    new_file_name = self.add_unique_filename(file_s_name + suffix)
                    network.name = new_file_name
                    self.s_data[new_file_name] = network
                    if hasattr(network, 'port_names') and network.port_names is not None:
                        print(f"处理后的端口名称：")
                        print(*network.port_names, sep='\n')
                except Exception as e:
                    QMessageBox.warning(self, '处理错误', f'文件 {file_name} 处理失败: {str(e)}')
                    continue
        except Exception as e:
            show_error(self, f"操作执行出错: {str(e)}")

    def call_port_reorder(self):
        try:
            selected_files = [item.text() for item in self.file_list.selectedItems()]
            if not selected_files:
                QMessageBox.warning(self, '错误', '请先选择文件！')
                return
            if len(selected_files) > 1:
                QMessageBox.information(self, "提示", "检测到多文件选择，请确定选择的S参数以同样的配置处理端口")

            file = selected_files[0]
            network = self.get_network(file)

            port_names = None
            if not network.port_names:
                dialog = PortNameDialog(self, network.nports, selected_files)
                port_names = dialog.get_port_names()
                if not port_names:
                    return
            else:
                port_names = network.port_names

            editor = PortOrderEditor(port_names, self)
            if editor.exec() == QDialog.DialogCode.Accepted:
                new_order_1based = editor.get_ordered_ports()
                print('新的端口顺序(1-based):', new_order_1based)
                new_order_0based = [x - 1 for x in new_order_1based]

                if len(new_order_0based) != network.nports:
                    QMessageBox.critical(self, "错误",
                                         f"端口数量不匹配: 新顺序有{len(new_order_0based)}个端口，"
                                         f"但应有{network.nports}个端口")
                    return
                if set(new_order_0based) != set(range(network.nports)):
                    QMessageBox.critical(self, "错误",
                                         f"端口索引不匹配:\n新顺序包含: {set(new_order_0based)}\n"
                                         f"但应有: {set(range(network.nports))}")
                    return

                for file_name in selected_files:
                    network_ori = self.get_network(file_name)
                    network = network_ori.copy()
                    if not network.port_names:
                        network.port_names = port_names
                    name_ori = network.name
                    network.renumber(new_order_0based, list(range(network.nports)))
                    print(f"端口重排完成~重新排布后的端口名称：")
                    print(*network.port_names, sep='\n')
                    suffix = f'_reorder.s{network.nports}p'
                    new_file_name = self.add_unique_filename(name_ori + suffix)
                    network.name = new_file_name
                    self.s_data[new_file_name] = network

                QMessageBox.information(self, "成功", f"已成功处理{len(selected_files)}个文件的端口重排序！")
        except Exception as e:
            show_error(self, f"操作执行出错: {str(e)}")
            traceback.print_exc()

    def call_port_merge(self):
        try:
            selected_files = [item.text() for item in self.file_list.selectedItems()]
            if not selected_files:
                QMessageBox.warning(self, '错误', '请先选择文件！')
                return
            if len(selected_files) > 1:
                QMessageBox.information(self, "提示",
                                        "检测到多文件选择，请确定选择的S参数以同样的配置处理端口")

            dialog = PortMergeDialog(self, selected_files)
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return

            merge_groups_1based, z0_list = dialog.get_result()
            print('端口合并配置:')
            for i, (group, z0) in enumerate(zip(merge_groups_1based, z0_list)):
                print(f'  第{i + 1}组: 端口 {group} → Merge_port, Z0={z0}Ω')

            for file_name in selected_files:
                try:
                    network_ori = self.get_network(file_name)
                    network = network_ori.copy()
                    n_ports = network.nports

                    merge_groups_0based = [[p - 1 for p in g] for g in merge_groups_1based]
                    all_ports = [p for g in merge_groups_0based for p in g]
                    if all_ports and max(all_ports) >= n_ports:
                        QMessageBox.warning(self, '错误',
                                            f'文件 {network.name} 端口范围超出！'
                                            f'最大端口数为 {n_ports}')
                        continue

                    new_network = merge_ports_multi(network, merge_groups_0based, z0_list)
                    suffix = f'_merged.s{new_network.nports}p'
                    base = os.path.splitext(network_ori.name)[0]
                    new_file_name = self.add_unique_filename(base + suffix)
                    new_network.name = new_file_name
                    self.s_data[new_file_name] = new_network
                    print(f'合并完成: {new_file_name}')
                    print('新端口名称:')
                    print(*new_network.port_names, sep='\n')
                except Exception as e:
                    QMessageBox.warning(self, '处理错误',
                                        f'文件 {file_name} 处理失败: {str(e)}')
                    continue
        except Exception as e:
            show_error(self, f"操作执行出错: {str(e)}")

    def call_port_management(self):
        dialog = PortManagementDialog(self)
        result = dialog.exec()
        dispatch = {
            1: self.call_edit_port_names,
            2: self.call_edit_z0,
            3: self.call_port_reorder,
            4: self.call_port_merge,
            5: self.call_port_reduction,
        }
        if result in dispatch:
            dispatch[result]()

    def call_edit_port_names(self):
        selected_files = [item.text() for item in self.file_list.selectedItems()]
        if not selected_files:
            QMessageBox.warning(self, '错误', '请先选择文件！')
            return
        if len(selected_files) > 1:
            QMessageBox.information(self, "提示", "仅编辑第一个选中文件的端口名，其余文件不受影响")
        file_name = selected_files[0]
        network = self.get_network(file_name)
        current_names = (network.port_names if network.port_names
                         else [f"Port{i + 1}" for i in range(network.nports)])
        prefill = "\n".join(current_names)
        helper = PortNameDialog(self, network.nports, file_name)
        new_names = helper._show_edit_dialog(prefill_text=prefill)
        if new_names:
            network.port_names = new_names
            self.s_param.pop(file_name, None)
            self.y_param.pop(file_name, None)
            self.z_param.pop(file_name, None)
            print(f"端口名称已更新（{file_name}）:")
            for i, name in enumerate(new_names, 1):
                print(f"  Port {i}: {name}")

    def call_edit_z0(self):
        selected_files = [item.text() for item in self.file_list.selectedItems()]
        if not selected_files:
            QMessageBox.warning(self, '错误', '请先选择文件！')
            return
        if len(selected_files) > 1:
            QMessageBox.information(self, "提示", "仅修改第一个选中文件的参考阻抗，其余文件不受影响")
        file_name = selected_files[0]
        network = self.get_network(file_name)
        dialog = Z0EditDialog(self, network)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_z0 = dialog.get_z0_values()
            if new_z0 is None:
                return
            for i, z0_val in enumerate(new_z0):
                network.z0[:, i] = z0_val
            self.s_param.pop(file_name, None)
            self.y_param.pop(file_name, None)
            self.z_param.pop(file_name, None)
            print(f"参考阻抗已更新（{file_name}）:")
            for i, z in enumerate(new_z0, 1):
                print(f"  Port {i}: {z:.1f} Ω")

    def call_cascade(self):
        try:
            selected_files = [item.text() for item in self.file_list.selectedItems()]
            if not selected_files:
                QMessageBox.warning(self, '错误', '请先选择文件！')
                return

            dialog = CascadeDialog(self, selected_files)
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return

            cascade_configs = dialog.get_result()
            networks = []
            for cfg in cascade_configs:
                file_name = cfg["sparam_file"]
                left_ports = cfg["ports_left"]
                right_ports = cfg["ports_right"]

                ntwk = self.get_network(file_name)
                enforce_nonzero_impedance(ntwk)

                total_ports = list(range(ntwk.nports))
                left_indices = [p - 1 for p in left_ports]
                right_indices = [p - 1 for p in right_ports]
                used_indices = left_indices + right_indices
                remaining_indices = [p for p in total_ports if p not in used_indices]
                new_order = left_indices + right_indices + remaining_indices
                new_ports = list(range(len(new_order)))
                ntwk_reordered = ntwk.copy()
                ntwk_reordered.renumber(new_order, new_ports)
                networks.append({
                    "ntwk": ntwk_reordered,
                    "left_count": len(left_ports),
                    "right_count": len(right_ports)
                })

            f_mins = [n["ntwk"].f[0] for n in networks]
            f_maxs = [n["ntwk"].f[-1] for n in networks]
            df_list = [np.mean(np.diff(n["ntwk"].f)) for n in networks]
            f_min = max(f_mins)
            f_max = min(f_maxs)
            df_common = min(df_list)
            f_new = np.arange(f_min, f_max, df_common)
            new_freq = rf.Frequency.from_f(f_new, unit='Hz')

            for n in networks:
                n["ntwk"] = n["ntwk"].interpolate(new_freq)

            result = networks[0]["ntwk"]
            name_first = result.name
            for i in range(1, len(networks)):
                left = result
                right = networks[i]["ntwk"]
                ports = networks[i - 1]["right_count"]
                result = rf.connect(left, left.nports - ports, right, 0, ports)

            suffix = f'_cascade.s{result.nports}p'
            new_file_name = self.add_unique_filename(name_first + suffix)
            result.name = new_file_name
            self.s_data[new_file_name] = result
            print(f'freq: {len(result.f)}')
            print(f'sparam: {len(result.s)}')
        except Exception as e:
            show_error(self, f"出错: {str(e)}")

    def call_ripple_dialog(self):
        selected_files = [item.text() for item in self.file_list.selectedItems()]
        if not selected_files:
            QMessageBox.warning(self, '错误', '请先选择文件！')
            return
        dialog = RippleFitDialog(self, selected_files)
        dialog.exec()

    def call_frequency_analysis_dialog(self):
        dialog = frequencyAnalysisDialog(self.s_data, self)
        dialog.show()

    def call_time_domain_dialog(self):
        from dialogs.time_domain import TimeDomainDialog
        dialog = TimeDomainDialog(self.s_data, self)
        dialog.show()
