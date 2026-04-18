# 建议保留的实际使用的导入
from datetime import datetime
import getpass
import subprocess

from PyQt6.QtGui import QTextCursor
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QFileDialog, QListWidget, QLabel, QLineEdit, QMessageBox,
    QComboBox, QGroupBox, QCheckBox, QGridLayout, QSpinBox,
    QTextEdit, QProgressDialog  # QListWidgetItem, QFrame 可删
)
from PyQt6.QtCore import Qt  # QObject, pyqtSignal, QEvent 未用可删
from openpyxl import Workbook, load_workbook
import matplotlib
from Basic_function_module import *
from UI2_Cascade import SParamCascadeDialog
from UI2_Frequency_Analysis import frequencyAnalysisDialog
from UI2_SE2Diff import DiffConversionDialog
from UI2_Port_reduction import PortReductionDialog
from UI2_PortOrderEditor import PortOrderEditor
from UI2_PortSelection import PortSelector
from portname_setting import PortNameDialog
from Longtime_block_hint import LoadingDialog


class SParameterViewer_MainWin(QWidget):
    def __init__(self):
        super().__init__()
        # 记录软件启动时间
        self.start_time = datetime.now()
        self.user_name = getpass.getuser()
        self.initUI()
        # 新增状态保存变量
        self.plot_history = []  # 保存所有绘图记录
        self.current_plot_data = []  # 当前待绘制的数据
        self.s_data = {}
        self.s_param = {}
        self.y_param = {}
        self.z_param = {}
        # 初始化
        self.loading = LoadingDialog(self)



    def initUI(self):
        # 主界面
        self.version_num = 'B2026.1'
        self.setWindowTitle(f'Quick_Sparam_{self.version_num} 封装SIPI开发部'
                            ' --- 本工具免费提供给其他组织使用，但对出现的问题、结果等概不负责')
        self.setGeometry(100, 100, 1500, 700)

        # 总布局（上下结构，住布局+信息输出框）
        # whole_layout = QVBoxLayout()
        # 主布局（左右结构）
        main_layout = QHBoxLayout()
        main_layout.setSpacing(15)  # 设置左右间距

        # ========== 左侧按钮栏 ==========
        # region 左侧按钮栏
        left_panel = QVBoxLayout()
        left_panel.setContentsMargins(5, 5, 15, 5)
        left_panel.setSpacing(10)  # 按钮间距

        # 文件操作按钮
        self.open_button = QPushButton('打开S参数')
        self.open_button.setFixedSize(120, 40)  # 统一按钮宽度
        self.open_button.clicked.connect(self.open_file_dialog)
        left_panel.addWidget(self.open_button)
        self.save_button = QPushButton('保存S参数')
        self.save_button.setFixedSize(120, 40)  # 统一按钮宽度
        self.save_button.clicked.connect(self.save_sparameters)
        left_panel.addWidget(self.save_button)
        self.read_button = QPushButton('查看源文件')
        self.read_button.setFixedSize(120, 40)  # 统一按钮宽度
        self.read_button.clicked.connect(self.read_snp_file)
        left_panel.addWidget(self.read_button)
        self.delete_button = QPushButton('删除S参数')
        self.delete_button.setFixedSize(120, 40)
        self.delete_button.clicked.connect(self.delete_selected_sparameters)
        left_panel.addWidget(self.delete_button)

        # 端口操作按钮——————简并、重排、级联
        self.port_reduction_button = QPushButton('端口reduction')
        self.port_reduction_button.setFixedSize(120, 40)
        self.port_reduction_button.clicked.connect(self.call_port_reduction)
        left_panel.addWidget(self.port_reduction_button)
        self.port_reorder_button = QPushButton('端口重排')
        self.port_reorder_button.setFixedSize(120, 40)
        self.port_reorder_button.clicked.connect(self.call_port_reorder)
        left_panel.addWidget(self.port_reorder_button)
        self.cascade_button = QPushButton('S参数级联')
        self.cascade_button.setFixedSize(120, 40)
        self.cascade_button.clicked.connect(self.call_cascade)
        left_panel.addWidget(self.cascade_button)

        # 差分转换按钮
        self.diff_button = QPushButton('差分转换')
        self.diff_button.setFixedSize(120, 40)
        self.diff_button.clicked.connect(self.call_diff_conversion)
        left_panel.addWidget(self.diff_button)

        # 频域分析按钮
        self.analysis_btn = QPushButton('频域分析')
        self.analysis_btn.setFixedSize(120, 40)
        self.analysis_btn.clicked.connect(self.call_frequency_analysis_dialog)
        left_panel.addWidget(self.analysis_btn)

        # 添加弹簧使按钮顶部对齐
        # left_panel.addStretch()

        # ========== 右侧内容区 ==========
        right_panel = QVBoxLayout()
        right_panel.setContentsMargins(5, 5, 5, 5)

        # 文件列表（占据右侧上半部分）
        file_group = QGroupBox("S参数文件列表")
        file_layout = QVBoxLayout()
        self.file_list = QListWidget()
        self.file_list.setSelectionMode(QListWidget.SelectionMode.ExtendedSelection)
        self.file_list.setMinimumWidth(400)  # 最小宽度保证可见性
        file_layout.addWidget(self.file_list)
        file_group.setLayout(file_layout)
        right_panel.addWidget(file_group, stretch=4)  # 4份高度
        # endregion

        # ========== 右侧控制&输出栏 ==========

        # region 绘图模块
        plot_group = QGroupBox("绘图控制")
        plot_layout = QHBoxLayout()  # 主布局改为水平
        # 端口输入组
        plot_left_layout = QGridLayout()
        plot_left_layout.addWidget(QLabel("端口1:"), 0, 0)
        self.port1_input = QLineEdit()
        self.port1_input.setPlaceholderText("例: 1 2 3 或 1:5")
        plot_left_layout.addWidget(self.port1_input, 0, 1, 1, 2)

        plot_left_layout.addWidget(QLabel("端口2:"), 1, 0)
        self.port2_input = QLineEdit()
        self.port2_input.setPlaceholderText("例: 1:2:5")
        plot_left_layout.addWidget(self.port2_input, 1, 1, 1, 2)
        self.plot_button = QPushButton('绘 图')
        self.plot_button.setFixedHeight(50)
        self.plot_button.clicked.connect(self.plot_s_parameters)
        plot_left_layout.addWidget(self.plot_button, 3, 0, 2, 3)

        # ===== 右侧区域 (其他选项) =====
        plot_right_layout = QGridLayout()
        self.port_select_btn = QPushButton("选择端口")
        self.port_select_btn.clicked.connect(self.on_port_select)
        self.port_select_btn.setFixedWidth(80)
        plot_right_layout.addWidget(self.port_select_btn, 0, 0)
        self.Basic_info = QPushButton('信息一览')
        self.Basic_info.clicked.connect(self.Basic_info_print)
        plot_right_layout.addWidget(self.Basic_info, 0, 2)
        self.BandWidth_focus = QPushButton('频段切片')
        self.BandWidth_focus.clicked.connect(self.BandWidth_focus_plot)
        plot_right_layout.addWidget(self.BandWidth_focus, 0, 3)

        plot_right_layout.addWidget(QLabel("映射模式:"), 1, 0)
        self.mapping_combo = QComboBox()
        self.mapping_combo.addItems(["一 一对应", "交叉映射"])
        plot_right_layout.addWidget(self.mapping_combo, 1, 1)

        plot_right_layout.addWidget(QLabel("数据模式:"), 2, 0)
        self.data_mode_combo = QComboBox()
        self.data_mode_combo.addItems([
            "幅度 (dB)", "幅度 (abs)", "相位 (度)", "相位 (rad)",
            "unwrap相位 (度)", "unwrap相位 (rad)", "群延迟 (fs)", "阻抗参数(mohm)", "导纳参数"
        ])
        plot_right_layout.addWidget(self.data_mode_combo, 2, 1)
        self.legend_checkbox = QCheckBox("显示图例")
        self.legend_checkbox.setChecked(True)
        plot_right_layout.addWidget(self.legend_checkbox, 3, 0)
        self.same_plot_checkbox = QCheckBox("曲线叠加")
        plot_right_layout.addWidget(self.same_plot_checkbox, 3, 1)
        plot_right_layout.addWidget(QLabel("关注频点:"), 3, 2)
        self.freG_input = QLineEdit("")  # 30
        self.freG_input.setPlaceholderText("GHz")
        plot_right_layout.addWidget(self.freG_input, 3, 3)


        plot_layout.addLayout(plot_left_layout, stretch=2)
        plot_layout.addLayout(plot_right_layout, stretch=2)
        plot_group.setLayout(plot_layout)

        # region Ripple分析模块
        fit_group = QGroupBox("Ripple拟合选项")
        fit_layout = QVBoxLayout()  # 改为垂直布局

        # 第一行：频率输入控件（横向排列）
        freq_input_layout = QHBoxLayout()
        freq_input_layout.addWidget(QLabel('起始频率 (GHz):'))
        self.start_freq_input = QLineEdit("0")
        # self.start_freq_input.setFixedWidth(80)  # 设置固定宽度
        freq_input_layout.addWidget(self.start_freq_input)
        freq_input_layout.addSpacing(10)  # 添加间距
        freq_input_layout.addWidget(QLabel('终止频率 (GHz):'))
        self.stop_freq_input = QLineEdit()
        # self.stop_freq_input.setFixedWidth(80)
        freq_input_layout.addWidget(self.stop_freq_input)
        freq_input_layout.addSpacing(10)
        freq_input_layout.addWidget(QLabel('频率间隔 (GHz):'))
        self.freq_step_input = QLineEdit("0")
        # self.freq_step_input.setFixedWidth(100)
        self.freq_step_input.setEnabled(False)  # 禁用控件
        freq_input_layout.addWidget(self.freq_step_input)

        freq_input_layout.addStretch()  # 添加伸缩项使控件左对齐
        fit_layout.addLayout(freq_input_layout)

        control_grid = QGridLayout()
        # 拟合方法选择
        control_grid.addWidget(QLabel("拟合方法:"), 1, 2)
        self.fit_method = QComboBox()
        self.fit_method.addItems(["n次多项式", "IEEE_std_802.3-2022", "平滑函数"])
        self.fit_method.currentTextChanged.connect(self._update_fit_ui)
        control_grid.addWidget(self.fit_method, 1, 3)
        # 多项式阶数输入
        self.poly_order_label = QLabel("多项式阶数:")
        self.poly_order_input = QSpinBox()
        self.poly_order_input.setRange(1, 10)
        self.poly_order_input.setValue(5)  # 默认值5
        control_grid.addWidget(self.poly_order_label, 2, 2)
        control_grid.addWidget(self.poly_order_input, 2, 3)
        # IEEE标准说明标签
        self.ieee_label = QLabel("93A-51公式: a0 + a1*sqrt(f) + a2*f + a4*f^2")
        self.ieee_label.setStyleSheet("color: #666; font-style: Times New Roman")
        self.ieee_label.hide()  # 默认隐藏
        control_grid.addWidget(self.ieee_label, 2, 2, 1, 2)
        self.ieee_label.setFixedSize(300, 20)
        # 平滑函数说明标签
        # 平滑函数参数输入控件
        self.smooth_window_label = QLabel("平滑窗长度:")
        self.smooth_window_input = QSpinBox()
        self.smooth_window_input.setRange(3, 51)  # 窗长度必须是奇数，最小3，最大51
        self.smooth_window_input.setValue(21)  # 默认值21
        self.smooth_window_input.setSingleStep(2)  # 步长2以保证奇数

        self.smooth_order_label = QLabel("阶数:")
        self.smooth_order_input = QSpinBox()
        self.smooth_order_input.setRange(1, 5)  # 阶数范围1-5
        self.smooth_order_input.setValue(3)  # 默认值3

        # 将控件添加到布局（第2行，跨2列）
        control_grid.addWidget(self.smooth_window_label, 2, 2)
        control_grid.addWidget(self.smooth_window_input, 2, 3)
        control_grid.addWidget(self.smooth_order_label, 3, 2)
        control_grid.addWidget(self.smooth_order_input, 3, 3)

        # 默认隐藏平滑函数控件
        self.smooth_window_label.hide()
        self.smooth_window_input.hide()
        self.smooth_order_label.hide()
        self.smooth_order_input.hide()
        # ripple分析按钮
        self.ripple_button = QPushButton('Ripple 分析')
        self.ripple_button.setFixedHeight(50)
        self.ripple_button.clicked.connect(self.call_ripple_analysis)
        control_grid.addWidget(self.ripple_button, 1, 0, 2, 2)
        fit_layout.addLayout(control_grid)
        fit_group.setLayout(fit_layout)



        # endregion


        right_panel.addWidget(plot_group, stretch=2)  # 2份高度
        right_panel.addWidget(fit_group, stretch=2)
        # ========== 组合布局 ==========
        main_layout.addLayout(left_panel, stretch=1)  # 左侧按钮栏占1份宽度
        main_layout.addLayout(right_panel, stretch=4)  # 右侧内容区占4份宽度

        # region log信息窗格
        # 创建输出文本框元素分组
        down_panel = QVBoxLayout()  # 主要为了标注分组
        output_group = QGroupBox("信息输出")
        # 上下结构，上文本下按键
        output_layout = QVBoxLayout()
        # 文本框只有一个元素，无所谓H V
        output_text_layout = QHBoxLayout()
        self.output_console = QTextEdit()
        self.output_console.setReadOnly(True)  # 设置为只读
        # self.output_console.setMinimumHeight(100)  # 设置最小高度
        output_text_layout.addWidget(self.output_console)
        # 左右结构一字排开
        output_button_layout = QHBoxLayout()  # 按键垂直分布
        clear_cache_button = QPushButton("清除缓存")
        clear_cache_button.clicked.connect(self.clear_cache_data)
        output_button_layout.addWidget(clear_cache_button)
        clear_button = QPushButton("清除输出")
        clear_button.clicked.connect(self.output_console.clear)
        output_button_layout.addWidget(clear_button)
        save_button = QPushButton("保存输出")
        save_button.clicked.connect(self.save_output_to_file)
        output_button_layout.addWidget(save_button)
        info_button = QPushButton("版本信息")
        info_button.clicked.connect(self.info_version)
        output_button_layout.addWidget(info_button)

        # 组合文本框分组
        output_layout.addLayout(output_text_layout)
        output_layout.addLayout(output_button_layout)
        output_group.setLayout(output_layout)
        down_panel.addWidget(output_group)
        # 重定向标准输出
        self._original_stdout = sys.stdout
        sys.stdout = self  # 将标准输出重定向到当前类
        main_layout.addLayout(down_panel, stretch=3)
        # endregion
        self.setLayout(main_layout)
        self._setup_ui_style()
# region UI支撑类函数

    def _setup_ui_style(self):
        """统一界面样式"""
        # 按钮样式
        button_style = """
        QPushButton {
            font-size: 14px;
            padding: 8px;
            border-radius: 5px;
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                      stop:0 #f6f7fa, stop:1 #dadbde);
            min-width: 80px;
        }
        QPushButton:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                      stop:0 #e7e8eb, stop:1 #cbcccf);
        }
        """

        # 应用样式
        for btn in [self.open_button, self.save_button, self.diff_button, self.port_reduction_button,
                    self.cascade_button, self.port_reorder_button, self.delete_button,
                    self.analysis_btn, self.plot_button, self.ripple_button, self.read_button]:
            btn.setStyleSheet(button_style)

        # 列表样式
        self.file_list.setStyleSheet("""
            QListWidget {
                font-size: 13px;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
            QListWidget::item {
                padding: 5px;
            }
        """)

    def _update_fit_ui(self, method):
        """根据选择的拟合方法更新UI"""
        # 隐藏所有参数控件
        self.poly_order_label.hide()
        self.poly_order_input.hide()
        self.ieee_label.hide()
        self.smooth_window_label.hide()
        self.smooth_window_input.hide()
        self.smooth_order_label.hide()
        self.smooth_order_input.hide()

        # 显示当前方法对应的控件
        if method == "n次多项式":
            self.poly_order_label.show()
            self.poly_order_input.show()
        elif method == "IEEE_std_802.3-2022":
            self.ieee_label.show()
        elif method == "平滑函数":
            self.smooth_window_label.show()
            self.smooth_window_input.show()
            self.smooth_order_label.show()
            self.smooth_order_input.show()

    def write(self, text):
        """重写write方法以捕获self.write输出"""
        cursor = self.output_console.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)  # 修改这里
        cursor.insertText(text)
        self.output_console.setTextCursor(cursor)
        self.output_console.ensureCursorVisible()

        # 同时保留原始输出到控制台
        # self._original_stdout.write(text)

    def check_beta_period(self):
        # 设置内测截止时间(2025年6月30日)
        beta_end_date = datetime(2026, 6, 30)
        current_date = datetime.now()

        if current_date > beta_end_date:
            QMessageBox.critical(
                self,
                "内测结束",
                "内测已结束，请联系管理员获取公测版本。",  # 内容
                QMessageBox.StandardButton.Ok  # 按钮
            )

            return False
        return True


    def closeEvent(self, event):
        """重写窗口关闭事件，自动写入使用日志"""
        self.on_app_closing()
        super().closeEvent(event)

    def on_app_closing(self):
        """软件关闭时触发的槽函数"""
        close_time = datetime.now()
        usage_duration = str(close_time - self.start_time)

        data_to_write = [
            self.user_name,
            self.start_time.strftime("%Y-%m-%d"),
            usage_duration
        ]

        print("写入使用日志:", data_to_write)
        # shared_folder = r"\\10.114.193.143\Public"
        if sys.platform == 'win32':
            pass
            # self.write_usage_to_network_excel(
            #     shared_folder=r"\\100.102.194.237\w00810255",
            #     filename="Quick_Sparam_usage_log.xlsx",
            #     data=data_to_write
            # )
        else:
            self.write_usage_to_network_excel(
                shared_folder=r"/data/Storage_pisi/w00810255/Qs",
                filename="Quick_Sparam_usage_log.xlsx",
                data=data_to_write
            )

    def write_usage_to_network_excel(self, shared_folder, filename, data):
        """将使用数据写入网络共享的Excel文件"""
        unc_path = os.path.join(shared_folder, filename)

        try:
            if os.path.exists(unc_path):
                wb = load_workbook(unc_path)
                ws = wb.active
                if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
                    ws.append(["用户名", "日期", "使用时长"])
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(["用户名", "日期", "使用时长"])

            ws.append(data)
            wb.save(unc_path)
            print(f"✅ 成功写入: {unc_path}")

        except PermissionError:
            print("❌ 无法写入文件，文件可能正在被其他用户打开。")
        except Exception as e:
            print(f"❌ 写入Excel失败: {e}")

    def clear_cache_data(self):
        """清除文件列表控件内容和缓存数据"""
        # 清空文件列表控件中的所有条目
        self.file_list.clear()

        # 清除与文件列表相关的缓存数据
        self.s_data.clear()

    def save_output_to_file(self):
        """将输出内容保存到文件"""
        # 获取当前输出文本
        output_text = self.output_console.toPlainText()
        if not output_text.strip():
            QMessageBox.warning(self, '无内容', '输出框中没有内容可保存！')
            return

        # 弹出文件保存对话框
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存输出内容",
            "",  # 起始目录为空表示使用默认目录
            "文本文件 (*.txt);;所有文件 (*)"
        )

        if file_path:
            try:
                # 确保文件以.txt结尾
                if not file_path.lower().endswith('.txt'):
                    file_path += '.txt'

                # 写入文件
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(output_text)

                # 显示成功消息
                QMessageBox.information(self, '保存成功', f'输出内容已保存到:\n{file_path}')
            except:
                show_error(self, "保存文件时出错")

    def info_version(self):
        print("B2026版主要更新内容如下：")
        items = [
            "---功能增加",
            "【频域分析】   功能中加入指定bit、最差bit的频域结果查看和比对",
            "【频域分析】   新增了VTF loss/crosstalk的数据类型",
            "【端口缩并】   新增了Cio端接的功能",
            "---交互体验",
            "【文件列表】   改为更为常用的Ctrl/Shift+点击的交互模式",
            "【端口重排】   支持端口批量选中、拖拽实现顺序调整",
            "【频域分析】   修复保存excel时文件名的长度限制",
            "【参数级联】   修复了特定级联后S参数保存的bug",
            "【参数读取】   HFSS生成的频变阻抗信息的S参数会导致报错，端口阻抗统一修复为标量阻抗",
        ]

        for text in items:
            print(f"{text}")
        print(f'\n版本持续迭代中，当前版本号为{self.version_num}')
# region 主界面功能

    def open_file_dialog(self):
        """弹出文件选择对话框，读取 S 参数文件并存储 Network 对象"""
        file_names, _ = QFileDialog.getOpenFileNames(
            self, '打开 S 参数文件', '', 'S 参数文件 (*.s*p *.S*P)'
        )
        if file_names:
            for file_name in file_names:
                self.file_list.addItem(file_name)

    def read_snp_file(self):
        selected_files = [item.text() for item in self.file_list.selectedItems()]
        for file in selected_files:
            if sys.platform == 'win32':
                subprocess.Popen(['notepad.exe', file])
            else:
                subprocess.Popen(['gvim', file])

    def save_sparameters(self):
        """保存选中的S参数文件到指定路径"""
        # 获取选中的文件
        selected_items = self.file_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, '警告', '请先在列表中选择要保存的S参数文件！')
            return

        # 弹出目录选择对话框
        save_dir = QFileDialog.getExistingDirectory(
            self,
            "选择保存目录",
            "",  # 默认目录（空表示当前目录）
            QFileDialog.Option.ShowDirsOnly
        )

        if not save_dir:  # 用户取消了选择
            return

        # 保存每个选中的文件
        success_count = 0
        for item in selected_items:
            try:
                file_path = item.text()
                network = get_network(self, file_path)

                # 构建新文件路径
                file_name = os.path.basename(file_path)
                save_path = os.path.join(save_dir, file_name)

                # 处理重名文件
                if os.path.exists(save_path):
                    base, ext = os.path.splitext(file_name)
                    counter = 1
                    while os.path.exists(os.path.join(save_dir, f"{base}_{counter}{ext}")):
                        counter += 1
                    save_path = os.path.join(save_dir, f"{base}_{counter}{ext}")

                # 保存文件
                # 判断端口阻抗是否一致
                z0_array = np.atleast_1d(network.z0)
                print(z0_array)
                all_equal_z0 = np.allclose(z0_array, z0_array[0, 0])
                print(all_equal_z0)

                if all_equal_z0:
                    try:
                        print('尝试使用默认方法保存（touchstone1.0）')
                        network.write_touchstone(save_path)
                    except UnicodeEncodeError:
                        # 即使阻抗一致也可能有编码问题，回退到手动保存
                        print('默认方法失败，尝试使用自定义方法保存（touchstone2.0）')
                        ts_string = network.write_touchstone(
                            return_string=True,
                            write_z0=True,
                        )
                        with open(save_path, 'w', encoding='utf-8') as f:
                            f.write(ts_string)
                else:
                    print("阻抗不一致，尝试使用自定义方法保存（touchstone2.0）")
                    # 阻抗不一致，使用 Touchstone 2.0 手动写入方式
                    ts_string = network.write_touchstone(
                        return_string=True,
                        write_z0=True,  # 写 Z0 注释
                    )
                    with open(save_path, 'w', encoding='utf-8') as f:
                        f.write(ts_string)
                success_count += 1

            except Exception as e:
                QMessageBox.warning(
                    self,
                    '保存失败',
                    f'文件 {os.path.basename(file_path)} 保存失败: {str(e)}'
                )

        # 显示保存结果
        if success_count > 0:
            QMessageBox.information(
                self,
                '保存成功',
                f'成功保存 {success_count}/{len(selected_items)} 个文件到:\n{save_dir}'
            )

    def delete_selected_sparameters(self):
        """删除选中的S参数文件和对应缓存"""
        selected_items = self.file_list.selectedItems()
        if not selected_items:
            pass

        for item in selected_items:
            file_name = item.text()
            # 从列表中移除
            self.file_list.takeItem(self.file_list.row(item))
            # 从缓存中删除
            if file_name in self.s_data:
                del self.s_data[file_name]


    def on_curve_click(self, event):
        """处理曲线点击事件"""

        if not event.artist in self.plot_lines:  # 使用存储的lines列表
            print('No line data')
            return

        line = event.artist
        self.highlight_curve(line)
        self.print_curve_info(line)
        self.fig.canvas.draw()  # 点击后刷新图形

    def highlight_curve(self, selected_line):
        """高亮显示选中的曲线"""
        for line in self.plot_lines:
            line.set_linewidth(1 if line != selected_line else 3)
            line.set_alpha(0.7 if line != selected_line else 1.0)

        # 高亮图例
        legend = self.ax.get_legend()
        if legend:
            for text in legend.get_texts():
                if text.get_text() == selected_line.get_label():
                    text.set_fontweight('bold')
                    text.set_color('red')
                else:
                    text.set_fontweight('normal')
                    text.set_color('black')

    def print_curve_info(self, line):
        """打印曲线信息"""
        info = f"""
        === 曲线信息 ===
        文件名: {line.network_name}
        参数: S{line.port_pair[0]},{line.port_pair[1]}
        模式: {line.data_mode}
        """
        # 频率范围: {line.freq_data[0]:.1f} - {line.freq_data[-1]:.1f} GHz
        # 数值范围: {np.min(line.value_data):.3f} - {np.max(line.value_data):.3f}
        print(info)

    def _plot_single_curve(self, network, szy_params, p1, p2):
        """绘制单条S参数曲线(内部辅助函数)"""
        # 检查端口是否有效
        num_port = network.number_of_ports
        if p1 > num_port or p2 > num_port:
            QMessageBox.warning(self, '端口错误',
                                f'文件 {network.name} 的端口{p1}或{p2}超出范围！')
            return
        # 获取当前数据模式
        data_mode = self.data_mode_combo.currentText()
        param = szy_params[:, p1 - 1, p2 - 1]  # 转换为0-based索引
        freqG = network.frequency.f / 1e9  # 将频率转换为GHz
        label = f'{network.name}_S{p1},{p2}'
        # print(data_mode)


        # 根据模式处理数据
        if data_mode == "幅度 (dB)":
            y_data = 20 * np.log10(np.abs(param))
        elif data_mode == "幅度 (abs)" or data_mode == "导纳参数":
            y_data = np.abs(np.abs(param))
        elif data_mode == "阻抗参数(mohm)":
            y_data = 1000 * np.abs(np.abs(param))
        elif data_mode == "相位 (度)":
            y_data = np.angle(param) * 180 / np.pi  # 转为角度
        elif data_mode == "相位 (rad)":
            y_data = np.angle(param)
        elif data_mode == "unwrap相位 (度)":
            y_data = np.unwrap(np.angle(param)) * 180 / np.pi
        elif data_mode == "unwrap相位 (rad)":
            y_data = np.unwrap(np.angle(param))
        elif data_mode == "群延迟 (fs)":
            phase = np.unwrap(np.angle(param))
            tau_g = -np.gradient(phase, freqG * 1e9) / (2 * np.pi)  # 秒
            y_data = tau_g * 1e12  # 转为飞秒

        # 绘制曲线并获取Line2D对象
        if self.legend_checkbox.isChecked():
            line, = self.ax.plot(freqG, y_data, label=label, picker=5)
        else:
            line, = self.ax.plot(freqG, y_data, picker=5)
        # 存储曲线信息
        line.network_name = network.name
        line.port_pair = (p1, p2)
        line.data_mode = data_mode
        line.freq_data = freqG
        line.value_data = y_data
        if data_mode == "阻抗参数(mohm)" or data_mode == "导纳参数":

            self.ax.set_xscale('log')
            self.ax.set_yscale('log')
        # =============== 绘制数据标注 ==================
        # 获取用户输入的标注频点
        input = self.freG_input.text().split()
        mark_freqGs = list(map(float, input))

        if mark_freqGs:
            # 绘制虚线
            for fG_mark in mark_freqGs:
                if min(freqG) <= fG_mark  <= max(freqG):
                    idx = np.abs(freqG - fG_mark).argmin()
                    actual_freq = freqG[idx]
                    actual_value = y_data[idx]
                    # 绘制垂直线
                    self.ax.axvline(x=actual_freq, linestyle=':', alpha=1)

                    # 绘制频段内的极值点
                    mask = (freqG <= fG_mark)
                    band_freqG = freqG[mask]
                    band_data = y_data[mask]
                    # 查找最大&最小值
                    max_idx = np.argmax(band_data)
                    min_idx = np.argmin(band_data)
                    max_freqG = band_freqG[max_idx]
                    max_value = band_data[max_idx]
                    min_freqG = band_freqG[min_idx]
                    min_value = band_data[min_idx]
                    # 用红点标注极值
                    self.ax.plot(max_freqG, max_value, 'ro', markersize=5)
                    self.ax.plot(min_freqG, min_value, 'ro', markersize=5)
                    # 打印到控制台
                    print(f"S{p1},{p2} {data_mode} 信息提取--- \n"
                               f"频率范围: {band_freqG[0]:.1f}-{band_freqG[-1]:.1f} GHz\n"
                               f"频率——max: {max_freqG:.3f}, min: {min_freqG:.3f}, end: {actual_freq:.3f}\n"
                               f"数值——max: {max_value:.3f}; min: {min_value:.3f}, end: {actual_value:.3f} \n")
                else:
                    print('关注频点超出现有的频率数据，无法进行数据统计')
        return line  # 返回line对象以便后续处理

    def plot_s_parameters(self):
        # 获取选中的文件/端口输入/映射模式
        selected_files = [item.text() for item in self.file_list.selectedItems()]
        port1 = self.port1_input.text().strip()
        port2 = self.port2_input.text().strip()
        mapping_mode = self.mapping_combo.currentText()

        # 检查输入是否有效
        if not selected_files or not port1 or not port2:
            QMessageBox.warning(self, '输入错误', '请在上方列表中选择文件! 并输入端口1和端口2！')
            return

        # 解析端口输入
        port1_list = parse_port_input(port1)
        port2_list = parse_port_input(port2)
        if port1_list is None or port2_list is None:
            return

        # 保存当前绘图请求
        new_plot_data = {
            'files': selected_files,
            'port1': port1_list,
            'port2': port2_list,
            'mode': mapping_mode
        }

        # 根据复选框决定是否保留历史
        if not self.same_plot_checkbox.isChecked():
            self.plot_history = [new_plot_data]  # 新建记录
        else:
            self.plot_history.append(new_plot_data)  # 追加记录

        # 建立图形，设置图形属性
        self.fig, self.ax = plt.subplots()
        if sys.platform == 'win32':
            plt.rcParams['font.sans-serif'] = ['SimHei']
        else:
            plt.rcParams['font.sans-serif'] = ['WenQuanYi Zen Hei']
        plt.rcParams['axes.unicode_minus'] = False
        data_mode = self.data_mode_combo.currentText()
        self.ax.set_ylabel(data_mode)
        self.ax.set_xlabel("频率 (GHz)")
        self.ax.grid(True)

        # 显示等待提示
        self.loading = LoadingDialog(self)
        self.loading.show()
        QApplication.processEvents()

        try:
            # 存储所有曲线对象
            self.plot_lines = []
            # 绘制每个文件的S参数
            for plot_data in self.plot_history:
                for i, file_name in enumerate(plot_data['files']):
                    if self.loading.cancelled:
                        break
                    # 更新进度信息（不影响进度条动画）
                    show_str = os.path.basename(file_name)
                    self.loading.set_message(f"正在处理文件: {show_str}")
                    QApplication.processEvents()

                    try:
                        network = get_network(self, file_name)
                        port1_plot = plot_data['port1']
                        port2_plot = plot_data['port2']
                        print(f"{file_name}")

                        if mapping_mode == "一 一对应":
                            if len(port1_plot) != len(port2_plot):
                                QMessageBox.warning(self, '输入错误', '一一对应模式需要端口数量相同！')
                                return

                            for p1, p2 in zip(port1_plot, port2_plot):
                                if data_mode == '阻抗参数(mohm)':
                                    z_params = get_z(self, file_name)
                                    line = self._plot_single_curve(network, z_params, p1, p2)
                                elif data_mode == '导纳参数':
                                    y_params = get_y(self, file_name)
                                    line = self._plot_single_curve(network, y_params, p1, p2)
                                else:
                                    s_params = get_s(self, file_name)
                                    line = self._plot_single_curve(network, s_params, p1, p2)
                                self.plot_lines.append(line)
                        elif mapping_mode == "交叉映射":
                            for p1 in port1_plot:
                                for p2 in port2_plot:
                                    if data_mode == '阻抗参数(mohm)':
                                        z_params = get_z(self, file_name)
                                        line = self._plot_single_curve(network, z_params, p1, p2)
                                    elif data_mode == '导纳参数':
                                        y_params = get_y(self, file_name)
                                        line = self._plot_single_curve(network, y_params, p1, p2)
                                    else:
                                        s_params = get_s(self, file_name)
                                        line = self._plot_single_curve(network, s_params, p1, p2)
                                    self.plot_lines.append(line)
                    except Exception as e:
                        show_error(self, f"处理文件 {file_name} 时出错: {str(e)}")
                        continue

            # 显示图形并设置交互
            self.ax.legend()
            if hasattr(self, '_cid'):
                self.fig.canvas.mpl_disconnect(self._cid)
            self._cid = self.fig.canvas.mpl_connect('pick_event', self.on_curve_click)
            self.fig.canvas.draw()  # 强制重绘
            self.fig.show()

        except Exception as e:
            show_error(self, f"绘图时遇到错误: {str(e)}")
        finally:
            if hasattr(self, 'loading'):
                self.loading.close()


    def on_port_select(self):
        """处理端口选择按钮点击，仅处理目标输入框设置"""
        try:
            target_input = self.port2_input if self.port2_input.hasFocus() else self.port1_input

            file_list = [a.text() for a in self.file_list.selectedItems()]
            selected_ports = check_and_set_port_names(self, file_list)
            if selected_ports:
                target_input.setText(" ".join(map(str, selected_ports)))

        except Exception as e:
            show_error(self, f"端口选择出错: {str(e)}")

    def BandWidth_focus_plot(self):
        print('该功能尚未开发')


    def Basic_info_print(self):
        """打印选中的S参数基本信息"""
        # 获取选中的文件
        try:
            selected_items = self.file_list.selectedItems()
            for item in selected_items:
                file_name = item.text()
                S = get_network(self, file_name)
                z0 = S.z0[0, :]
                freq = S.f/1e9
                num_port = S.number_of_ports
                portnames = S.port_names
                print(f'文件名：{file_name}')
                print(f'频率范围：{freq[0]} ~ {freq[-1]} GHz')
                print(f'频点数：{len(freq)}')
                print(f'端口数量：{num_port}')

                # print(portnames)
                if portnames is not None and len(portnames) > 0:
                    print('端口名称：')
                    for i, (port_name, zref) in enumerate(zip(portnames, abs(z0)), start=1):
                        print(f'Port {i}: {port_name},   Z_ref={int(zref)}')
                else:
                    print(f'该S参数没有可用的端口名称信息')

        except:
            show_error(self, "打印信息时出错")

# endregion

    # region 次级界面功能


    def call_diff_conversion(self):
        selected_files = [item.text() for item in self.file_list.selectedItems()]
        if not selected_files:
            QMessageBox.warning(self, '错误', '请先选择文件！')
            return

        dialog = DiffConversionDialog(len(selected_files[0].split('.')[0].split('s')[-1][0]))  # 从文件名获取端口数
        if dialog.exec() == QDialog.DialogCode.Accepted:
            try:
                params = dialog.get_conversion_params()
                if not params:
                    return

                print("转换参数:")
                [print(f"{k} = {v}") for k, v in params.items()]

                z0_diff = params['z0_diff']
                port_mode = params['port_mode']
                output_mode = params['output_mode']
                logic_mode = params['Logic']

                for file_name in selected_files:
                    try:
                        network = get_network(self, file_name)
                        if not network:
                            continue

                        # 根据逻辑模式选择处理方式
                        if logic_mode == 'line':
                            if 'diff_list' in params and params['diff_list']:  # 部分差分模式
                                line_list = params['diff_list']
                                new_network = SE2dq_dqs(network, line_list, port_mode, z0_diff)
                            else:  # 全端口差分模式
                                new_network = SE2diff(network, port_mode, output_mode, z0_diff)
                        else:  # 端口逻辑模式
                            # 端口模式 - z0_diff是单值
                            z0_diff = params['z0_diff']  # 单个阻抗值
                            new_network = SE2diff_port(network, params['diff_list'],
                                                       z0_diff, params['output_mode'])

                        if new_network:
                            new_name = add_unique_filename(self, new_network.name)
                            self.s_data[new_name] = new_network
                            if new_network.port_names:
                                print('转换后的端口名称:')
                                print(*new_network.port_names, sep="\n")

                    except Exception as e:
                        QMessageBox.critical(self, "转换错误",
                                             f"处理文件 {file_name} 时出错:\n{str(e)}")
                        traceback.print_exc()

            except Exception as e:
                QMessageBox.critical(self, "参数错误", f"获取参数时出错:\n{str(e)}")
                traceback.print_exc()

    def call_port_reduction(self):
        """处理端口操作（支持RC复合端接）"""
        try:
            selected_files = [item.text() for item in self.file_list.selectedItems()]
            if not selected_files:
                QMessageBox.warning(self, '错误', '请先选择文件！')
                return
            if len(selected_files) > 1:
                QMessageBox.information(self, "提示", "检测到多文件选择，请确定选择的S参数以同样的配置处理端口")

            # 弹出设置对话框（新版支持RC设置）
            dialog = PortReductionDialog(self, selected_files)
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return

            # 获取配置数据（z0_configs现在是(R,C)元组列表）
            port_configs, z0_configs, port_delete = dialog.get_result()
            print('端口设置详情')
            print(f'端口组：{port_configs}')
            print(f'端接配置(R,C)：{z0_configs}')
            print(f'简并端口：{port_delete}')

            for file_name in selected_files:
                try:
                    network_ori = get_network(self, file_name)
                    enforce_nonzero_impedance(network_ori)
                    network = network_ori.copy()
                    n_ports = network.nports
                    freq = network.frequency.f  # 获取频率点(Hz)
                    file_s_name = network.name

                    # 验证端口范围
                    all_ports = [p for group in port_configs for p in group]
                    if all_ports and max(all_ports) > n_ports:
                        QMessageBox.warning(self, '错误',
                                            f'文件 {file_s_name} 的端口范围超出！最大端口数为 {n_ports}')
                        continue

                    if port_delete and max(port_delete) > n_ports:
                        QMessageBox.warning(self, '错误',
                                            f'文件 {file_s_name} 要删除的端口超出范围！最大端口数为 {n_ports}')
                        continue

                    # 创建新的参考阻抗数组
                    z0_new = network.z0.copy()

                    # 应用阻抗设置（每组独立计算）
                    for ports, z_rc in zip(port_configs, z0_configs):
                        port_indices = [p - 1 for p in ports]
                        Z_load = dialog.compute_load_impedance(freq, z_rc)
                        for p in port_indices:
                            z0_new[:, p] = Z_load

                    # 执行重归一化
                    network.renormalize(z0_new)
                    # ====================================

                    # 端口删除处理
                    if port_delete:
                        keep_ports = [i for i in range(n_ports) if (i + 1) not in port_delete]
                        network = network.subnetwork(ports=keep_ports)

                    # 生成新文件名并保存
                    # 判断是否所有电容值都为0
                    all_c_zero = all(c == 0 for (_, c) in z0_configs)
                    # 根据电容值选择后缀
                    if all_c_zero:
                        suffix = f'_renorm_R.s{network.nports}p'  # 纯电阻情况
                    else:
                        suffix = f'_renorm_RC.s{network.nports}p'  # 包含电容的情况

                    new_file_name = add_unique_filename(self, file_s_name + suffix)
                    network.name = new_file_name
                    self.s_data[new_file_name] = network
                    if hasattr(network, 'port_names') and network.port_names is not None:
                        print(f"处理后的端口名称：")
                        print(*network.port_names, sep='\n')


                except Exception as e:
                    QMessageBox.warning(self, '处理错误',
                                        f'文件 {file_name} 处理失败: {str(e)}')
                    continue

        except Exception as e:
            show_error(self, f"操作执行出错: {str(e)}")

    def call_port_reorder(self):
        """调用端口顺序编辑器并应用新的端口顺序"""
        try:
            # 获取当前网络的端口信息
            selected_files = [item.text() for item in self.file_list.selectedItems()]
            if not selected_files:
                QMessageBox.warning(self, '错误', '请先选择文件！')
                return

            if len(selected_files) > 1:
                QMessageBox.information(self, "提示", "检测到多文件选择，请确定选择的S参数以同样的配置处理端口")

            file = selected_files[0]
            network = get_network(self, file)

            # 处理没有端口名称的情况
            # 处理没有端口名称的情况
            port_names = None
            if not network.port_names:
                dialog = PortNameDialog(self, network.nports, selected_files)
                port_names = dialog.get_port_names()
                if not port_names:  # 如果返回 None 或空列表
                    return  # 用户取消操作
            else:
                port_names = network.port_names

            # 创建并显示端口顺序编辑器
            editor = PortOrderEditor(port_names, self)
            if editor.exec() == QDialog.DialogCode.Accepted:
                # 获取新的端口顺序 (1-based)
                new_order_1based = editor.get_ordered_ports()
                print('新的端口顺序(1-based):', new_order_1based)

                # 转换为0-based索引
                new_order_0based = [x - 1 for x in new_order_1based]

                # 验证新顺序的有效性
                if len(new_order_0based) != network.nports:
                    QMessageBox.critical(self, "错误",
                                         f"端口数量不匹配: 新顺序有{len(new_order_0based)}个端口，"
                                         f"但应有{network.nports}个端口")
                    return

                if set(new_order_0based) != set(range(network.nports)):
                    QMessageBox.critical(self, "错误",
                                         f"端口索引不匹配:\n新顺序包含: {set(new_order_0based)}\n"
                                         f"但应有: {set(range(network.nports))}")
                    return

                for file_name in selected_files:
                    network_ori = get_network(self, file_name)
                    network = network_ori.copy()

                    # 如果没有端口名称，使用用户提供的名称
                    if not network.port_names:
                        network.port_names = port_names

                    name_ori = network.name

                    # 执行端口重排序
                    network.renumber(new_order_0based, list(range(network.nports)))
                    print(f"端口重排完成~重新排布后的端口名称：")
                    print(*network.port_names, sep='\n')

                    # 生成新文件名并保存
                    suffix = f'_reorder.s{network.nports}p'
                    new_file_name = add_unique_filename(self, name_ori + suffix)
                    network.name = new_file_name
                    self.s_data[new_file_name] = network

                QMessageBox.information(self, "成功", f"已成功处理{len(selected_files)}个文件的端口重排序！")

        except Exception as e:
            show_error(self, f"操作执行出错: {str(e)}")
            import traceback
            traceback.print_exc()  # 打印完整错误堆栈以便调试


    def call_cascade(self):
        # print('该功能初步完成，尚未完全测试，欢迎试用反馈')
        try:
            selected_files = [item.text() for item in self.file_list.selectedItems()]
            if not selected_files:
                QMessageBox.warning(self, '错误', '请先选择文件！')
                return

            # 弹出设置对话框
            dialog = SParamCascadeDialog(self, selected_files)
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return

            cascade_configs = dialog.get_result()

            networks = []
            for cfg in cascade_configs:
                file_name = cfg["sparam_file"]
                left_ports = cfg["ports_left"]
                right_ports = cfg["ports_right"]

                # 获取网络对象
                ntwk = get_network(self, file_name)
                enforce_nonzero_impedance(ntwk)

                # 当前所有端口编号（从 0 开始）
                total_ports = list(range(ntwk.nports))

                # 将用户指定的 left/right 端口转为 0-based 索引
                left_indices = [p - 1 for p in left_ports]
                right_indices = [p - 1 for p in right_ports]

                # 已使用的端口
                used_indices = left_indices + right_indices
                # 剩余未参与连接的端口（保持原顺序）
                remaining_indices = [p for p in total_ports if p not in used_indices]
                # 构造完整的新顺序
                new_order = left_indices + right_indices + remaining_indices
                # 目标顺序下，输出端口编号应为 0~N-1
                new_ports = list(range(len(new_order)))
                # 创建新网络并重排端口顺序
                ntwk_reordered = ntwk.copy()
                ntwk_reordered.renumber(new_order, new_ports)

                networks.append({
                    "ntwk": ntwk_reordered,
                    "left_count": len(left_ports),
                    "right_count": len(right_ports)
                })

            # 确定公共频率范围并插值
            f_mins = [n["ntwk"].f[0] for n in networks]
            f_maxs = [n["ntwk"].f[-1] for n in networks]
            df_list = [np.mean(np.diff(n["ntwk"].f)) for n in networks]

            f_min = max(f_mins)
            f_max = min(f_maxs)
            df_common = min(df_list)
            f_new = np.arange(f_min, f_max, df_common)
            new_freq = rf.Frequency.from_f(f_new, unit='Hz')

            # 插值所有网络
            for n in networks:
                n["ntwk"] = n["ntwk"].interpolate(new_freq)

            # 执行级联（从第一个开始，依次 connect）
            result = networks[0]["ntwk"]
            name_first = result.name
            for i in range(1, len(networks)):
                left = result
                right = networks[i]["ntwk"]
                ports = networks[i - 1]["right_count"]  # 上一个网络的右端口数量
                result = rf.connect(left, left.nports - ports, right, 0, ports)

            # 展示结果（示例：保存或绘图）
            suffix = f'_cascade.s{result.nports}p'
            new_file_name = add_unique_filename(self, name_first + suffix)
            result.name = new_file_name
            self.s_data[new_file_name] = result
            print(f'freq: {len(result.f)}')
            print(f'sparam: {len(result.s)}')

        except Exception as e:
            show_error(self, f"出错: {str(e)}")

    def call_ripple_analysis(self):
        try:
            # 获取UI控件输入
            selected_files = [item.text() for item in self.file_list.selectedItems()]
            port1 = self.port1_input.text().strip()
            port2 = self.port2_input.text().strip()
            start_freq = self.start_freq_input.text().strip()
            stop_freq = self.stop_freq_input.text().strip()
            method = self.fit_method.currentText()
            order = self.poly_order_input.value() if method == "n次多项式" else 0
            data_mode = self.data_mode_combo.currentText()

            # 输入验证
            if not (selected_files and port1 and port2 and start_freq and stop_freq):
                QMessageBox.warning(self, '输入错误', '请选择文件并输入端口1、端口2、起始频率和终止频率！')
                return

            try:
                port1_list = parse_port_input(port1)
                port2_list = parse_port_input(port2)
                start_freq = float(start_freq)
                stop_freq = float(stop_freq)
            except (ValueError, TypeError) as e:
                QMessageBox.warning(self, '输入错误', f'输入参数格式错误: {str(e)}')
                return

            # 处理数据
            results_data = []
            results_text = []
            for file_name in selected_files:
                try:
                    network = get_network(self, file_name)
                    for p1, p2 in zip(port1_list, port2_list):
                        if max(p1, p2) > network.s.shape[1]:
                            QMessageBox.warning(self, '端口错误', f'文件 {file_name} 的端口数不足！')
                            continue
                        if method == "n次多项式":
                            fit_params = {'order': self.poly_order_input.value()}
                        elif method == "平滑函数":
                            fit_params = {
                                'window_length': self.smooth_window_input.value(),
                                'polyorder': self.smooth_order_input.value()
                            }
                        else:  # IEEE标准
                            fit_params = {}

                        result = ripple_calc(network, p1, p2, start_freq, stop_freq,
                                             data_mode, method, fit_params)
                        results_data.append(result)
                        results_text.append(f"{result['label']}: ripple = {result['max_ripple']:.4f} ")
                except Exception as e:
                    show_error(self, f"处理文件 {file_name} 时出错: {str(e)}")
                    continue

            # 绘制图形
            if results_data:
                plot_main_curves(results_data, data_mode)
                plot_residuals(results_data, data_mode)

                # 显示结果文本
                print('\nRipple 分析结果:')
                for text in results_text:
                    print(text)
            else:
                QMessageBox.warning(self, '无结果', '没有生成任何有效结果！')
        except:
            show_error(self, "保存文件时出错")

    def call_frequency_analysis_dialog(self):
        # 获取选中的文件
        # selected_files = [item.text() for item in self.file_list.selectedItems()]
        # if not selected_files:
        #     QMessageBox.warning(self, '错误', '请先选择S参数文件！')
        #     return

        dialog = frequencyAnalysisDialog(self.s_data, self)
        dialog.show()

        # endregion


    # endregion

